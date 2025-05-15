from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.generic import FormView, ListView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q, Sum, Avg, Min, Max, Count
from django.utils.dateparse import parse_datetime, parse_date
import json
from .forms import ReportCreationForm
from .models import SavedReport
from django.urls import reverse, reverse_lazy
from django.core.management.base import BaseCommand
from django.db.models import ForeignKey
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
import logging
from django.utils import timezone
from django.core.serializers.json import DjangoJSONEncoder

# Create your views here.

logger = logging.getLogger(__name__)

class ReportCreationView(LoginRequiredMixin, FormView):
    template_name = 'reporting/report_creation.html'
    form_class = ReportCreationForm
    success_url = reverse_lazy('reporting:report_list')
    
    def setup(self, request, *args, **kwargs):
        """Initialize attributes shared by all view methods."""
        super().setup(request, *args, **kwargs)
        self.report_id = kwargs.get('report_id')
        self.report = None
        if self.report_id:
            self.report = get_object_or_404(
                SavedReport, 
                id=self.report_id,
                user=request.user
            )
    
    def dispatch(self, request, *args, **kwargs):
        """Add cache control headers to prevent browser caching."""
        response = super().dispatch(request, *args, **kwargs)
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response
    
    def get_initial(self):
        """Provide initial form data for editing."""
        initial = super().get_initial()
        if self.report:
            # Log the raw data
            print("Raw report data:", {
                'aggregations': self.report.aggregations,
                'type': type(self.report.aggregations)
            })
            
            # Ensure all data is properly serialized
            initial.update({
                'report_name': self.report.name,
                'selected_tables': json.dumps(self.report.selected_tables or [], cls=DjangoJSONEncoder),
                'selected_fields': json.dumps(self.report.selected_fields or {}, cls=DjangoJSONEncoder),
                'filters': json.dumps(self.report.filters or [], cls=DjangoJSONEncoder),
                'sort_by': json.dumps(self.report.sort_by or {}, cls=DjangoJSONEncoder),
                'sort_direction': self.report.sort_direction or 'asc',
                'aggregations': json.dumps(self.report.aggregations or {}, cls=DjangoJSONEncoder),
                'group_by': json.dumps(self.report.group_by or {}, cls=DjangoJSONEncoder)
            })
            
            # Log the processed data
            print("Processed initial data:", initial)
        return initial
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        form = context['form']
        
        # Add available tables to context
        context['available_tables'] = [
            {
                'name': model_key,
                'verbose_name': model_class._meta.verbose_name.title()
            }
            for model_key, model_class in form.AVAILABLE_MODELS.items()
        ]
        
        # Add edit context
        context['is_edit'] = bool(self.report)
        context['report'] = self.report
        
        # Log initial data for debugging
        if self.report:
            print("Report Data:", {
                'id': self.report.id,
                'name': self.report.name,
                'selected_tables': self.report.selected_tables,
                'selected_fields': self.report.selected_fields,
                'aggregations': self.report.aggregations
            })
        
        return context
    
    def form_valid(self, form):
        cleaned = form.cleaned_data
        
        try:
            # Parse and validate selected_tables
            selected_tables = cleaned['selected_tables']
            if isinstance(selected_tables, str):
                selected_tables = json.loads(selected_tables)
            if not isinstance(selected_tables, list):
                selected_tables = list(selected_tables)
                
            # Parse and validate selected_fields
            selected_fields = cleaned['selected_fields']
            if isinstance(selected_fields, str):
                selected_fields = json.loads(selected_fields)
            if not isinstance(selected_fields, dict):
                raise ValueError("Selected fields must be a dictionary")
                
            # Parse and validate filters
            filters = cleaned['filters']
            if isinstance(filters, str):
                filters = json.loads(filters)
            if not isinstance(filters, list):
                filters = []
                
            # Validate each filter object
            validated_filters = []
            for filter_item in filters:
                if isinstance(filter_item, dict) and all(k in filter_item for k in ['table', 'field', 'operator', 'value']):
                    validated_filters.append({
                        'table': str(filter_item['table']),
                        'field': str(filter_item['field']),
                        'operator': str(filter_item['operator']),
                        'value': filter_item['value']
                    })

            # Parse and validate sort configuration
            sort_by = cleaned.get('sort_by', '{}')
            sort_direction = cleaned.get('sort_direction', 'asc')
            
            if isinstance(sort_by, str):
                sort_by = json.loads(sort_by)
            if not isinstance(sort_by, dict):
                sort_by = {}
                
            # Validate sort configuration
            if sort_by:
                for table, config in sort_by.items():
                    if not isinstance(config, dict) or 'field' not in config:
                        sort_by = {}
                        break
                    if table not in selected_tables:
                        sort_by = {}
                        break
            
            # Ensure sort_direction is valid
            if sort_direction not in ['asc', 'desc']:
                sort_direction = 'asc'
            
            # Parse and validate aggregations with better error handling
            aggregations = cleaned.get('aggregations', '{}')
            print("Raw aggregations data:", aggregations)  # Debug log
            
            if isinstance(aggregations, str):
                try:
                    aggregations = json.loads(aggregations)
                except json.JSONDecodeError as e:
                    print(f"Error decoding aggregations JSON: {e}")
                    aggregations = {}
            
            if not isinstance(aggregations, dict):
                print(f"Invalid aggregations type: {type(aggregations)}")
                aggregations = {}
            
            # Validate each aggregation configuration
            validated_aggregations = {}
            for field_path, config in aggregations.items():
                if isinstance(config, dict) and 'type' in config:
                    if config['type'] in ['sum', 'avg', 'min', 'max', 'count']:
                        validated_aggregations[field_path] = {
                            'type': config['type'],
                            'label': config.get('label', '')
                        }
            
            # Parse and validate group by with better error handling
            group_by = cleaned.get('group_by', '{}')
            print("Raw group_by data:", group_by)  # Debug log
            
            if isinstance(group_by, str):
                try:
                    group_by = json.loads(group_by)
                except json.JSONDecodeError as e:
                    print(f"Error decoding group_by JSON: {e}")
                    group_by = {}
            
            if not isinstance(group_by, dict):
                print(f"Invalid group_by type: {type(group_by)}")
                group_by = {}
            
            # Validate group by configuration
            validated_group_by = {}
            for table, fields in group_by.items():
                if isinstance(fields, list):
                    validated_group_by[table] = [str(field) for field in fields]
            
            print("Final validated data:")  # Debug log
            print("Aggregations:", validated_aggregations)
            print("Group By:", validated_group_by)
            
            report_data = {
                'name': cleaned['report_name'],
                'selected_tables': selected_tables,
                'selected_fields': selected_fields,
                'filters': validated_filters,
                'sort_by': sort_by,
                'sort_direction': sort_direction,
                'aggregations': validated_aggregations,
                'group_by': validated_group_by
            }
            
            if self.report:
                # Update existing report
                for key, value in report_data.items():
                    setattr(self.report, key, value)
                self.report.save()
            else:
                # Create new report
                report_data['user'] = self.request.user
                SavedReport.objects.create(**report_data)
            
            return super().form_valid(form)
            
        except Exception as e:
            print(f"Error in form_valid: {e}")  # Debug log
            form.add_error(None, f"Error processing form data: {str(e)}")
            return self.form_invalid(form)

def get_model_fields(request):
    """API endpoint to get fields for selected tables."""
    try:
        selected_tables = json.loads(request.GET.get('selected_tables', '[]'))
        if not selected_tables:
            return JsonResponse({'error': 'No tables selected'}, status=400)
        
        form = ReportCreationForm()
        
        # Get any additional tables needed for relationships
        linking_tables = form.get_linking_tables(selected_tables)
        if linking_tables:
            all_tables = selected_tables + linking_tables
            message = f"Added linking tables: {', '.join(linking_tables)}"
        else:
            all_tables = selected_tables
            message = None
        
        # Get available fields for all tables
        fields = form.get_available_fields(all_tables)
        
        return JsonResponse({
            'fields': fields,
            'linking_tables': linking_tables,
            'message': message
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

def get_table_relationships(request):
    """API endpoint to get relationships between tables."""
    try:
        form = ReportCreationForm()
        return JsonResponse({
            'relationships': form.model_relationships
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

def get_field_values(request):
    """API endpoint to get unique values for a field (with optional search term)."""
    try:
        field_name = request.GET.get('field')
        search_term = request.GET.get('term', '')
        limit = int(request.GET.get('limit', 10))  # Default to 10 results
        
        if not field_name:
            return JsonResponse({'error': 'Field name is required'}, status=400)
            
        # Parse the field name (format: "table.field")
        try:
            table_name, field = field_name.split('.')
        except ValueError:
            return JsonResponse({'error': 'Invalid field name format'}, status=400)
            
        # Get the model class
        form = ReportCreationForm()
        if table_name not in form.AVAILABLE_MODELS:
            return JsonResponse({'error': 'Invalid table name'}, status=400)
            
        model_class = form.AVAILABLE_MODELS[table_name]
        
        # Build the query
        query = model_class.objects.all()
        if search_term:
            # Use icontains for case-insensitive search
            query = query.filter(**{f"{field}__icontains": search_term})
        
        # Get distinct values
        values = query.values_list(field, flat=True).distinct().order_by(field)[:limit]
        
        return JsonResponse({
            'values': list(values)
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

class SavedReportListView(LoginRequiredMixin, ListView):
    model = SavedReport
    template_name = 'reporting/report_list.html'
    context_object_name = 'reports'
    
    def get_queryset(self):
        return SavedReport.objects.filter(user=self.request.user)

class ReportDisplayView(LoginRequiredMixin, ListView):
    template_name = 'reporting/report_display.html'
    context_object_name = 'results'
    paginate_by = 25  # Show 25 items per page
    
    def setup(self, request, *args, **kwargs):
        """Initialize attributes shared by all view methods."""
        super().setup(request, *args, **kwargs)
        self.report = get_object_or_404(
            SavedReport, 
            id=kwargs.get('report_id'),
            user=request.user
        )
    
    def get_queryset(self):
        """
        Construct and return the queryset based on the saved report configuration.
        """
        if not self.report.selected_tables:
            return []
            
        primary_table = self.report.selected_tables[0]
        form = ReportCreationForm()
        
        if primary_table not in form.AVAILABLE_MODELS:
            return []
            
        model_class = form.AVAILABLE_MODELS[primary_table]
        queryset = model_class.objects.all()
        
        # Build relationship paths for all tables
        relationship_map = {}
        related_fields = set()
        
        # First, find all relationship paths
        for table in self.report.selected_tables:
            if table != primary_table:
                path = self._find_relationship_path(model_class, table, form.AVAILABLE_MODELS)
                if path:
                    relationship_map[table] = path
                    related_fields.add(path)
        
        # Apply select_related for all related tables
        if related_fields:
            queryset = queryset.select_related(*related_fields)
        
        # Apply filters
        filters = self.report.filters
        if filters:
            q_objects = Q()
            for filter_item in filters:
                if isinstance(filter_item, dict):
                    filter_table = filter_item.get('table')
                    field = filter_item.get('field')
                    operator = filter_item.get('operator', 'equals')
                    value = filter_item.get('value')
                    
                    if field and value is not None:
                        # Determine the full field path
                        if filter_table == primary_table:
                            field_path = field
                        elif filter_table in relationship_map:
                            field_path = f"{relationship_map[filter_table]}__{field}"
                        else:
                            continue
                        
                        filter_lookup = self._build_filter_lookup(field_path, operator, value)
                        if filter_lookup:
                            if filter_lookup.get('_negated'):
                                del filter_lookup['_negated']
                                q_objects &= ~Q(**filter_lookup)
                            else:
                                q_objects &= Q(**filter_lookup)
            
            if q_objects:
                queryset = queryset.filter(q_objects)
        
        # Apply sorting
        sort_by = self.report.sort_by
        if sort_by:
            sort_fields = []
            for table, config in sort_by.items():
                if isinstance(config, dict) and 'field' in config:
                    field = config['field']
                    if table == primary_table:
                        field_path = field
                    elif table in relationship_map:
                        field_path = f"{relationship_map[table]}__{field}"
                    else:
                        continue
                    
                    if self.report.sort_direction == 'desc':
                        field_path = f"-{field_path}"
                    sort_fields.append(field_path)
            
            if sort_fields:
                queryset = queryset.order_by(*sort_fields)
        
        # Handle group by and aggregations
        group_by = self.report.group_by
        aggregations = self.report.aggregations
        
        if group_by and isinstance(group_by, dict):
            # Build list of fields to group by
            group_fields = []
            for table, fields in group_by.items():
                if isinstance(fields, list):
                    for field in fields:
                        if table == primary_table:
                            group_fields.append(field)
                        elif table in relationship_map:
                            group_fields.append(f"{relationship_map[table]}__{field}")
            
            if group_fields and aggregations:
                # Build aggregation configuration
                aggregation_config = {}
                for field_path, config in aggregations.items():
                    if isinstance(config, dict) and 'type' in config:
                        agg_type = config['type']
                        if agg_type == 'sum':
                            aggregation_config[f"{field_path}__{agg_type}"] = Sum(field_path)
                        elif agg_type == 'avg':
                            aggregation_config[f"{field_path}__{agg_type}"] = Avg(field_path)
                        elif agg_type == 'min':
                            aggregation_config[f"{field_path}__{agg_type}"] = Min(field_path)
                        elif agg_type == 'max':
                            aggregation_config[f"{field_path}__{agg_type}"] = Max(field_path)
                        elif agg_type == 'count':
                            aggregation_config[f"{field_path}__{agg_type}"] = Count(field_path)
                
                # Apply group by with aggregations
                queryset = queryset.values(*group_fields).annotate(**aggregation_config)
                
                # Store group by fields for template
                self.group_fields = group_fields
                
                # Store aggregation info for template
                self.aggregation_info = {
                    field_path: config
                    for field_path, config in aggregations.items()
                }
                
                return queryset.order_by(*group_fields)
        
        # If no grouping, just return all fields
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Add group by and aggregation information to context
        if hasattr(self, 'group_fields'):
            context['group_fields'] = self.group_fields
        if hasattr(self, 'aggregation_info'):
            context['aggregation_info'] = self.aggregation_info
        
        # Add field labels
        field_labels = []
        primary_table = self.report.selected_tables[0]
        form = ReportCreationForm()
        
        if primary_table in form.AVAILABLE_MODELS:
            model_class = form.AVAILABLE_MODELS[primary_table]
            
            # Build relationship paths
            relationship_map = {}
            for table in self.report.selected_tables:
                if table != primary_table:
                    path = self._find_relationship_path(model_class, table, form.AVAILABLE_MODELS)
                    if path:
                        relationship_map[table] = path
            
            # Add field labels
            selected_fields = self.report.selected_fields
            if isinstance(selected_fields, dict):
                for table, fields in selected_fields.items():
                    if table in form.AVAILABLE_MODELS:
                        model_class = form.AVAILABLE_MODELS[table]
                        
                        for field_name in fields:
                            try:
                                field = model_class._meta.get_field(field_name)
                                # For related tables, prefix the field name with the relationship path
                                if table != primary_table and table in relationship_map:
                                    field_name = f"{relationship_map[table]}__{field_name}"
                                
                                field_labels.append({
                                    'name': field_name,
                                    'label': f"{model_class._meta.verbose_name.title()} {field.verbose_name.title()}"
                                })
                            except Exception as e:
                                print(f"Error getting field label for {field_name}: {e}")
                                field_labels.append({
                                    'name': field_name,
                                    'label': field_name.replace('_', ' ').title()
                                })
            
            context['field_labels'] = field_labels
        
        return context
    
    def _find_relationship_path(self, start_model, target_table, available_models, visited_tables=None):
        """
        Find the path of relationships from start_model to target_table.
        Returns the path as a string (e.g., 'clin__contract' or 'supplier__clin')
        """
        if visited_tables is None:
            visited_tables = set()
            
        if target_table not in available_models:
            return None
            
        target_model = available_models[target_table]
        
        # Direct relationship check
        for field in start_model._meta.get_fields():
            if hasattr(field, 'related_model') and field.related_model:
                if field.related_model == target_model:
                    return field.name
                    
        # Check for indirect relationships through other models
        for field in start_model._meta.get_fields():
            if hasattr(field, 'related_model') and field.related_model:
                intermediate_model = field.related_model
                intermediate_table = None
                
                # Find the table name for this model
                for table, model in available_models.items():
                    if model == intermediate_model:
                        intermediate_table = table
                        break
                
                if (intermediate_table and 
                    intermediate_table not in visited_tables and 
                    intermediate_table in self.report.selected_tables):
                    visited_tables.add(intermediate_table)
                    next_path = self._find_relationship_path(
                        intermediate_model, 
                        target_table, 
                        available_models,
                        visited_tables
                    )
                    if next_path:
                        return f"{field.name}__{next_path}"
        
        return None
    
    def _build_filter_lookup(self, field, operator, value):
        """
        Build the filter lookup dictionary based on the operator.
        Returns a dictionary with the appropriate Django ORM lookup.
        """
        # Initialize table and field variables
        table = None
        field_name = field
        
        # Handle field name with table prefix (e.g., 'contract.status' -> 'status')
        if '.' in field:
            table, field_name = field.split('.')
            
        # Get the model class and field object
        form = ReportCreationForm()
        field_obj = None
        
        if table and table in form.AVAILABLE_MODELS:
            model_class = form.AVAILABLE_MODELS[table]
            try:
                field_obj = model_class._meta.get_field(field_name)
                # If this is a choice field, we should use exact matching
                if hasattr(field_obj, 'choices') and field_obj.choices:
                    operator = 'equals'  # Force exact matching for choice fields
            except Exception as e:
                print(f"Error getting field object: {e}")
            
        operator_map = {
            'equals': {'lookup': '', 'value': value},
            'not_equals': {'lookup': '', 'value': value, 'negate': True},
            'contains': {'lookup': '__icontains', 'value': value},
            'not_contains': {'lookup': '__icontains', 'value': value, 'negate': True},
            'gt': {'lookup': '__gt', 'value': value},
            'gte': {'lookup': '__gte', 'value': value},
            'lt': {'lookup': '__lt', 'value': value},
            'lte': {'lookup': '__lte', 'value': value},
            'starts_with': {'lookup': '__istartswith', 'value': value},
            'ends_with': {'lookup': '__iendswith', 'value': value},
            'in': {'lookup': '__in', 'value': value.split(',') if isinstance(value, str) else value},
            'not_in': {'lookup': '__in', 'value': value.split(',') if isinstance(value, str) else value, 'negate': True},
            'is_null': {'lookup': '__isnull', 'value': True},
            'is_not_null': {'lookup': '__isnull', 'value': False},
        }
        
        if operator not in operator_map:
            return None
            
        op_config = operator_map[operator]
        lookup = f"{field_name}{op_config['lookup']}"
        
        if op_config.get('negate'):
            return {f"{lookup}": op_config['value'], '_negated': True}
        return {lookup: op_config['value']}

class ExportReportToExcelView(LoginRequiredMixin, ListView):
    template_name = 'reporting/report_display.html'
    context_object_name = 'results'
    
    def _prepare_value_for_excel(self, value):
        """Prepare value for Excel export, handling timezone-aware datetimes."""
        if isinstance(value, datetime):
            # Convert timezone-aware datetime to naive datetime
            if timezone.is_aware(value):
                return timezone.make_naive(value)
        return value
    
    def setup(self, request, *args, **kwargs):
        logger.debug("ExportReportToExcelView.setup called")
        super().setup(request, *args, **kwargs)
        self.report = get_object_or_404(
            SavedReport, 
            id=kwargs.get('report_id'),
            user=request.user
        )
        logger.debug(f"Report loaded: {self.report.id}")
    
    def get_queryset(self):
        logger.debug("ExportReportToExcelView.get_queryset called")
        report_view = ReportDisplayView()
        report_view.report = self.report
        report_view.request = self.request
        report_view.args = self.args
        report_view.kwargs = self.kwargs
        queryset = report_view.get_queryset()
        logger.debug(f"Queryset obtained with {queryset.count()} items")
        return queryset
    
    def get(self, request, *args, **kwargs):
        logger.debug("ExportReportToExcelView.get called")
        try:
            self.object_list = self.get_queryset()
            
            report_view = ReportDisplayView()
            report_view.report = self.report
            report_view.request = request
            report_view.args = args
            report_view.kwargs = kwargs
            report_view.object_list = self.object_list
            
            context_data = report_view.get_context_data()
            field_labels = context_data['field_labels']
            logger.debug(f"Got field labels: {len(field_labels)} fields")
            
            # Create a new workbook and select the active sheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Report Data"
            
            # Style for headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Write headers
            for col, field in enumerate(field_labels, start=1):
                cell = ws.cell(row=1, column=col, value=field['label'])
                cell.font = header_font
                cell.fill = header_fill
            
            # Write data
            for row_num, row in enumerate(self.object_list, start=2):
                for col_num, field in enumerate(field_labels, start=1):
                    value = row.get(field['name'], '')
                    # Prepare value for Excel
                    excel_value = self._prepare_value_for_excel(value)
                    ws.cell(row=row_num, column=col_num, value=excel_value)
            
            # Adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = min(adjusted_width, 50)  # Cap width at 50
            
            # Create response with Excel file
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{self.report.name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            
            # Save workbook to response
            wb.save(response)
            return response
            
        except Exception as e:
            logger.error(f"Error in ExportReportToExcelView.get: {str(e)}", exc_info=True)
            raise

class Command(BaseCommand):
    help = 'Fix malformed report data in the database'

    def handle(self, *args, **options):
        from reporting.models import SavedReport
        
        reports = SavedReport.objects.all()
        fixed_count = 0
        
        for report in reports:
            try:
                # Fix selected_tables
                if isinstance(report.selected_tables, str):
                    report.selected_tables = json.loads(report.selected_tables)
                if not isinstance(report.selected_tables, list):
                    report.selected_tables = list(report.selected_tables)
                
                # Fix selected_fields
                if isinstance(report.selected_fields, str):
                    report.selected_fields = json.loads(report.selected_fields)
                if not isinstance(report.selected_fields, dict):
                    # Try to convert to proper format
                    fixed_fields = {}
                    if isinstance(report.selected_fields, list):
                        # If it's a list of "table.field", convert to proper structure
                        for field in report.selected_fields:
                            if '.' in field:
                                table, field_name = field.split('.')
                                if table not in fixed_fields:
                                    fixed_fields[table] = []
                                fixed_fields[table].append(field_name)
                    report.selected_fields = fixed_fields
                
                # Fix filters
                if isinstance(report.filters, str):
                    report.filters = json.loads(report.filters)
                if not isinstance(report.filters, list):
                    report.filters = []
                
                # Validate and fix each filter
                fixed_filters = []
                for filter_item in report.filters:
                    if isinstance(filter_item, dict) and all(k in filter_item for k in ['table', 'field', 'operator', 'value']):
                        fixed_filters.append({
                            'table': str(filter_item['table']),
                            'field': str(filter_item['field']),
                            'operator': str(filter_item['operator']),
                            'value': filter_item['value']
                        })
                report.filters = fixed_filters
                
                report.save()
                fixed_count += 1
                self.stdout.write(self.style.SUCCESS(f'Successfully fixed report {report.id}'))
                
            except Exception as e:
                self.stdout.write(self.style.ERROR(f'Error fixing report {report.id}: {str(e)}'))
                
        self.stdout.write(self.style.SUCCESS(f'Fixed {fixed_count} reports'))
