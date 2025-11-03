from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from .forms import (
    UserRegisterForm,
    BaseForm,
    AdminLoginForm,
    PasswordChangeForm,
    PasswordSetForm,
    EmailLookupForm,
    OAuthPasswordSetForm,
    PortalSectionForm,
    PortalResourceForm,
    WorkCalendarEventForm,
    WorkCalendarTaskForm,
    EventAttachmentForm,
)
from django.contrib.auth.signals import user_logged_out
from django.dispatch import receiver
from STATZWeb.decorators import login_required
from django.http import JsonResponse, QueryDict
from django.contrib.auth.decorators import user_passes_test
from .models import (
    Announcement,
    AppPermission,
    CalendarAnalyticsSnapshot,
    EventAttendance,
    NaturalLanguageScheduleRequest,
    PortalResource,
    PortalSection,
    ScheduledMicroBreak,
    SystemMessage,
    UserSetting,
    UserSettingState,
    WorkCalendarEvent,
    WorkCalendarTask,
    EventAttachment,
    RecurrenceRule,
)
from contracts.models import Company
from django.contrib.auth.models import User
from django.urls import resolve, reverse
import logging
from django.views.decorators.http import require_http_methods
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import user_passes_test
import json
from .user_settings import UserSettings
from django.contrib.auth import get_user_model
from django.contrib.auth import authenticate, login as auth_login
from .ms_views import get_microsoft_login_url
from django.views.generic import ListView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.forms import AuthenticationForm
from django.conf import settings
from datetime import datetime, timedelta
from django.utils import timezone
from django.db import transaction
from django.db.models import Avg, Q
from .portal_services import (
    build_portal_context,
    get_visible_sections,
    serialize_resource,
    serialize_section,
    serialize_event,
    upcoming_events_for_user,
    active_tasks_for_user,
    serialize_task,
    upcoming_microbreaks,
    serialize_microbreak,
    outstanding_nlp_requests,
    serialize_nlp_request,
)

logger = logging.getLogger(__name__)


def _is_portal_admin(user):
    return bool(user and user.is_authenticated and (user.is_superuser or user.is_staff))


def _is_section_editor(user, section):
    if not user or not user.is_authenticated:
        return False
    if _is_portal_admin(user):
        return True
    return section.editors.filter(pk=user.pk).exists()


def _request_data(request):
    if request.content_type and 'application/json' in request.content_type:
        try:
            payload = json.loads(request.body.decode('utf-8'))
            return payload if isinstance(payload, dict) else {}
        except (ValueError, TypeError):
            return {}
    if request.method in ('GET', 'DELETE'):
        return request.GET
    return request.POST


def _as_int_list(value):
    if value is None:
        return []
    if isinstance(value, (list, tuple)):
        return [int(v) for v in value if str(v).isdigit()]
    if isinstance(value, str):
        parts = [item.strip() for item in value.split(',') if item.strip()]
        return [int(v) for v in parts if v.isdigit()]
    return []


try:
    from dateutil import parser as date_parser
except ImportError:  # pragma: no cover - defensive fallback
    date_parser = None


def _parse_natural_language_request(text, user):
    """
    Basic natural language parsing tailored to scheduling intents.
    Returns dict with success flag, proposed start/end, attendees, diagnostics.
    """
    lowered = text.lower()
    now = timezone.now()
    diagnostics = {'source': 'rule-based-v1'}

    # Duration detection (default 60 minutes)
    duration = 60
    import re

    duration_match = re.search(r'(\d+)\s*(minute|min|hour|hr)', lowered)
    if duration_match:
        value = int(duration_match.group(1))
        unit = duration_match.group(2)
        duration = value * 60 if unit.startswith('hour') or unit.startswith('hr') else value
    diagnostics['duration_detected'] = duration

    # Determine base date
    base_date = now
    if 'tomorrow' in lowered:
        base_date = now + timedelta(days=1)
    elif 'next week' in lowered:
        days_until_monday = (7 - now.weekday()) % 7 or 7
        base_date = now + timedelta(days=days_until_monday)
    elif 'next ' in lowered:
        days_lookup = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']
        for offset, day_name in enumerate(days_lookup):
            token = f'next {day_name}'
            if token in lowered:
                days_ahead = (offset - now.weekday()) % 7
                days_ahead = days_ahead + (7 if days_ahead == 0 else 0)
                base_date = now + timedelta(days=days_ahead)
                break
    else:
        days_lookup = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']
        for offset, day_name in enumerate(days_lookup):
            if day_name in lowered:
                days_ahead = (offset - now.weekday()) % 7
                base_date = now + timedelta(days=days_ahead)
                break

    # Time of day heuristics
    default_time = 9
    if 'before lunch' in lowered:
        default_time = 11
    elif 'after lunch' in lowered:
        default_time = 13
    elif 'lunch' in lowered:
        default_time = 12
    elif 'morning' in lowered:
        default_time = 9
    elif 'afternoon' in lowered:
        default_time = 14
    elif 'evening' in lowered:
        default_time = 18

    proposed_start = base_date.replace(hour=default_time, minute=0, second=0, microsecond=0)

    # Explicit time parsing
    time_match = re.search(r'(\d{1,2})(?::(\d{2}))?\s*(am|pm)?', lowered)
    if time_match:
        hour = int(time_match.group(1))
        minute = int(time_match.group(2) or 0)
        meridiem = time_match.group(3)
        if meridiem:
            if meridiem == 'pm' and hour < 12:
                hour += 12
            if meridiem == 'am' and hour == 12:
                hour = 0
        proposed_start = proposed_start.replace(hour=hour, minute=minute)
    elif date_parser:
        try:
            parsed_dt = date_parser.parse(text, fuzzy=True, default=now.replace(hour=default_time, minute=0, second=0, microsecond=0))
            if parsed_dt:
                proposed_start = timezone.make_aware(parsed_dt) if timezone.is_naive(parsed_dt) else parsed_dt
        except (ValueError, OverflowError):
            diagnostics['dateutil_parse'] = 'failed'

    proposed_end = proposed_start + timedelta(minutes=duration)

    # Attendee detection (naive)
    attendee_names = []
    name_match = re.search(r'with\s+([A-Z][a-zA-Z]+(?:[\s,&]+[A-Z][a-zA-Z]+)*)', text)
    if name_match:
        raw_names = re.split(r'[\s,&]+', name_match.group(1).strip())
        attendee_names = [name for name in raw_names if name]

    diagnostics['attendees_raw'] = attendee_names

    return {
        'success': True,
        'start': proposed_start,
        'end': proposed_end,
        'duration': duration,
        'attendees': attendee_names,
        'diagnostics': diagnostics,
    }


def _next_available_slot(user, proposed_start, duration_minutes, buffer_minutes=15):
    """
    Given a proposed start time, return a slot that avoids conflicts.
    """
    candidate_start = proposed_start
    candidate_end = candidate_start + timedelta(minutes=duration_minutes)
    conflicts = WorkCalendarEvent.objects.filter(
        organizer=user,
        start_at__lt=candidate_end,
        end_at__gt=candidate_start,
    ).order_by('start_at')

    for event in conflicts:
        candidate_start = event.end_at + timedelta(minutes=buffer_minutes)
        candidate_end = candidate_start + timedelta(minutes=duration_minutes)
    return candidate_start, candidate_end, bool(conflicts)


def _calculate_predicted_attendance(user, kind):
    aggregate = EventAttendance.objects.filter(
        event__organizer=user,
        event__kind=kind,
        confidence_score__isnull=False,
    ).aggregate(avg_conf=Avg('confidence_score'))
    avg_conf = aggregate.get('avg_conf')
    if avg_conf is None:
        return 0.75  # optimistic default until data accumulates
    return max(0.0, min(1.0, float(avg_conf)))


def _auto_insert_microbreak(user, event, duration_minutes=10):
    break_start = event.end_at
    break_end = break_start + timedelta(minutes=duration_minutes)
    has_conflict = WorkCalendarEvent.objects.filter(
        organizer=user,
        start_at__lt=break_end,
        end_at__gt=break_start,
    ).exists()
    if has_conflict:
        return None
    overlap_break = ScheduledMicroBreak.objects.filter(
        user=user,
        start_at__lt=break_end,
        end_at__gt=break_start,
    ).exists()
    if overlap_break:
        return None
    return ScheduledMicroBreak.objects.create(
        user=user,
        start_at=break_start,
        end_at=break_end,
        label='Recovery Break',
        insertion_mode='auto',
        related_event=event,
        notes='Auto-inserted after scheduling to prevent burnout.',
    )


def _parse_datetime_string(value):
    if not value:
        raise ValueError("No datetime provided")
    if date_parser:
        dt = date_parser.parse(value)
    else:
        dt = datetime.fromisoformat(value)
    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone.get_current_timezone())
    return dt
def login_view(request):
    """Custom login view that redirects to Microsoft authentication"""
    # Clear previous auth errors
    auth_error = None
    microsoft_auth_success = request.session.get('microsoft_auth_success', False)
    
    # Check for auth errors
    if 'microsoft_auth_error' in request.session:
        auth_error = request.session.pop('microsoft_auth_error')
        logger.warning(f"Microsoft auth error: {auth_error}")
    
    # Check if already authenticated
    if request.user.is_authenticated:
        logger.debug(f"User {request.user.username} is already authenticated")
        return redirect('index')
    
    # Get the next URL from query parameters
    next_url = request.GET.get('next')
    
    # Check for Microsoft auth success
    if microsoft_auth_success:
        logger.debug(f"Microsoft auth successful, redirecting to: {next_url or 'index'}")
        return redirect(next_url or 'index')
    
    # Get Microsoft login URL
    microsoft_login_url = get_microsoft_login_url(request)
    if next_url:
        microsoft_login_url = f"{reverse('users:microsoft_login')}?next={next_url}"
    
    # Handle password login form
    password_form = None
    
    if request.method == 'POST' and 'password_login' in request.POST:
        password_form = AdminLoginForm(data=request.POST)
        if password_form.is_valid():
            username = password_form.cleaned_data.get('username')
            password = password_form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                auth_login(request, user)
                messages.success(request, f'Welcome back, {user.username}!')
                return redirect(next_url or 'index')
            else:
                auth_error = 'Invalid username or password.'
        else:
            auth_error = 'Please correct the errors below.'
    else:
        password_form = AdminLoginForm()
    
    context = {
        'title': 'Login',
        'auth_error': auth_error,
        'microsoft_login_url': microsoft_login_url,
        'form': password_form,
    }
    
    return render(request, 'users/login.html', context)

def register(request):
    """Redirect registration to Microsoft authentication"""
    messages.info(request, 'New accounts are created through Microsoft authentication. Please sign in with Microsoft.')
    return redirect('users:microsoft_login')


@login_required
def profile(request):
    """User profile page showing account information and password management"""
    context = {
        'title': 'User Profile',
    }
    return render(request, 'users/profile.html', context)


@receiver(user_logged_out)
def on_user_logged_out(sender, request, **kwargs):
    messages.success(request, 'You have been successfully logged out.')

def permission_denied(request):
    return render(request, 'users/permission_denied.html')

@login_required
@require_POST
def switch_company(request):
    """Endpoint to switch the active company for the session.
    - Superusers can select any active company
    - Non-superusers must have membership in the selected company
    """
    next_url = request.POST.get('next') or request.META.get('HTTP_REFERER') or 'index'
    company_id = request.POST.get('company_id')
    if not company_id:
        messages.error(request, 'No company selected.')
        return redirect(next_url)
    try:
        company = Company.objects.get(pk=company_id, is_active=True)
    except Company.DoesNotExist:
        messages.error(request, 'Selected company not found or inactive.')
        return redirect(next_url)

    # Enforce membership for non-superusers
    if not request.user.is_superuser:
        from users.models import UserCompanyMembership
        allowed = UserCompanyMembership.objects.filter(user=request.user, company=company).exists()
        if not allowed:
            messages.error(request, 'You do not have access to the selected company.')
            return redirect(next_url)

    request.session['active_company_id'] = company.id
    # Persist selection in user settings for stability across sessions
    try:
        UserSettings.save_setting(
            user=request.user,
            name='current_company_id',
            value=company.id,
            setting_type='integer',
            description="User's currently selected company"
        )
    except Exception:
        pass
    messages.success(request, f'Active company set to {company.name}.')
    return redirect(next_url)

def is_staff(user):
    return user.is_staff

@user_passes_test(is_staff)
def debug_app_permissions(request):
    """A debug view to see all current app permissions in the database"""
    users = User.objects.all()
    
    # Build a dictionary of all permissions
    permissions_data = {}
    
    for user in users:
        permissions = AppPermission.objects.filter(user=user)
        user_permissions = {}
        
        for perm in permissions:
            user_permissions[perm.app_name_id] = perm.has_access
        
        permissions_data[user.username] = {
            'user_id': user.id,
            'permissions': user_permissions
        }
    
    return JsonResponse({
        'app_permissions': permissions_data,
        'total_users': users.count(),
        'total_permissions': AppPermission.objects.count()
    })

def test_app_name(request):
    """Test view to determine the current app_name"""
    resolved = resolve(request.path_info)
    app_name = resolved.app_name
    namespace = resolved.namespace
    url_name = resolved.url_name
    view_name = f"{namespace}:{url_name}" if namespace else url_name
    
    logger.info(f"app_name: {app_name}")
    logger.info(f"namespace: {namespace}")
    logger.info(f"url_name: {url_name}")
    logger.info(f"view_name: {view_name}")
    
    return JsonResponse({
        'app_name': app_name,
        'namespace': namespace,
        'url_name': url_name,
        'view_name': view_name,
        'path': request.path_info,
    })

@login_required
def save_user_setting(request):
    """Handle saving user settings via AJAX"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        data = json.loads(request.body)
        print(f"Received setting data: {data}")  # Debug print
        setting_type = data.get('setting_type')
        value = data.get('value')
        
        if not setting_type or value is None:
            print(f"Missing required fields: setting_type={setting_type}, value={value}")  # Debug print
            return JsonResponse({'success': False, 'error': 'Missing required fields'})
        
        success = UserSettings.save_setting(
            user=request.user,
            name=setting_type,
            value=value
        )
        
        print(f"Save result: {success}")  # Debug print
        return JsonResponse({
            'success': success,
            'message': 'Setting saved successfully' if success else 'Failed to save setting'
        })
        
    except json.JSONDecodeError:
        print("Invalid JSON data received")  # Debug print
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
    except Exception as e:
        print(f"Error saving setting: {str(e)}")  # Debug print
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def manage_settings(request):
    """View for managing user settings."""
    User = get_user_model()
    
    # Only get active users
    users = User.objects.filter(is_active=True)
    
    # Get all settings for the current user
    current_user_settings = UserSettings.get_all_settings(request.user)
    
    # Get all possible setting names from the database
    all_settings = UserSetting.objects.all().values_list('name', flat=True).distinct()
    
    # Add debug info
    logger.debug(f"Manage settings view: Found {users.count()} active users, {len(current_user_settings)} settings for current user, and {len(all_settings)} total unique settings")
    
    context = {
        'users': users,
        'current_user_settings': current_user_settings,
        'all_settings': list(all_settings),
        'setting_types': ['string', 'boolean', 'integer', 'float'],  # Available setting types
    }
    
    return render(request, 'users/manage_settings.html', context)

@login_required
@require_http_methods(["POST"])
def ajax_save_setting(request):
    """AJAX endpoint for saving user settings."""
    try:
        data = json.loads(request.body)
        user_id = data.get('user_id')
        setting_name = data.get('setting_name')
        setting_value = data.get('setting_value')
        setting_type = data.get('setting_type', 'string')
        
        User = get_user_model()
        user = User.objects.get(id=user_id)
        
        # Convert value based on setting type
        if setting_type == 'boolean':
            setting_value = setting_value.lower() == 'true'
        elif setting_type == 'integer':
            setting_value = int(setting_value)
        elif setting_type == 'float':
            setting_value = float(setting_value)
        
        success = UserSettings.save_setting(
            user=user,
            name=setting_name,
            value=setting_value,
            setting_type=setting_type
        )
        
        return JsonResponse({
            'success': success,
            'message': 'Setting saved successfully' if success else 'Failed to save setting'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)

@login_required
@require_http_methods(["GET"])
def ajax_get_user_setting(request):
    """AJAX endpoint for getting a user's settings."""
    try:
        user_id = request.GET.get('user_id')
        setting_name = request.GET.get('setting_name')
        
        logger.debug(f"ajax_get_user_setting called for user_id={user_id}, setting_name={setting_name}")
        
        if not user_id:
            return JsonResponse({
                'success': False,
                'message': 'User ID is required'
            }, status=400)
        
        User = get_user_model()
        user = User.objects.get(id=user_id)
        
        if setting_name:
            # Get specific setting
            value = UserSettings.get_setting(user, setting_name)
            logger.debug(f"Returning specific setting {setting_name}={value}")
            return JsonResponse({
                'success': True,
                'value': value
            })
        else:
            # Get all settings
            settings = UserSettings.get_all_settings(user)
            logger.debug(f"Returning all settings for user {user.username}: {settings}")
            return JsonResponse({
                'success': True,
                'settings': settings
            })
            
    except User.DoesNotExist:
        logger.error(f"User not found: {user_id}")
        return JsonResponse({
            'success': False,
            'message': 'User not found'
        }, status=404)
    except Exception as e:
        logger.error(f"Error getting user settings: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f"Error: {str(e)}"
        }, status=400)

@login_required
@require_http_methods(["GET"])
def ajax_get_setting_types(request):
    """AJAX endpoint for getting all setting types."""
    try:
        # Import here to avoid circular imports
        from .models import UserSetting
        
        # Get all settings with their types
        settings = UserSetting.objects.all()
        
        # Create dictionary of setting name to type
        types_dict = {setting.name: setting.setting_type for setting in settings}
        
        logger.debug(f"Returning {len(types_dict)} setting types")
        
        return JsonResponse({
            'success': True,
            'types': types_dict
        })
    except Exception as e:
        logger.error(f"Error getting setting types: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f"Error: {str(e)}"
        }, status=400)

@login_required
def check_auth_method(request):
    """
    Check and display the user's authentication method
    For debugging purposes
    """
    auth_method = request.session.get('auth_method', 'unknown')
    logger.info(f"User {request.user.username} authenticated via {auth_method}")
    
    # Check if microsoft token exists
    ms_token = request.session.get('microsoft_token', None)
    ms_token_status = "exists" if ms_token else "missing"
    
    return JsonResponse({
        'username': request.user.username,
        'email': request.user.email,
        'auth_method': auth_method,
        'microsoft_token_status': ms_token_status,
        'is_authenticated': request.user.is_authenticated,
    })

class SystemMessageListView(LoginRequiredMixin, ListView):
    """View for displaying all system messages for a user."""
    
    model = SystemMessage
    template_name = 'users/system_messages.html'
    context_object_name = 'messages'
    
    def get_queryset(self):
        return SystemMessage.objects.filter(user=self.request.user)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['unread_count'] = SystemMessage.get_unread_count(self.request.user)
        return context

class MarkMessageReadView(LoginRequiredMixin, View):
    """View for marking a message as read."""
    
    def post(self, request, *args, **kwargs):
        message_id = kwargs.get('pk')
        try:
            message = SystemMessage.objects.get(id=message_id, user=request.user)
            message.mark_as_read()
            return JsonResponse({'status': 'success'})
        except SystemMessage.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Message not found'}, status=404)

class MarkAllMessagesReadView(LoginRequiredMixin, View):
    """View for marking all messages as read."""
    
    def post(self, request, *args, **kwargs):
        messages = SystemMessage.objects.filter(user=request.user, read_at__isnull=True)
        for message in messages:
            message.mark_as_read()
        return JsonResponse({'status': 'success'})

class GetUnreadCountView(LoginRequiredMixin, View):
    """View for getting the count of unread messages."""
    
    def get(self, request, *args, **kwargs):
        count = SystemMessage.get_unread_count(request.user)
        return JsonResponse({'count': count})

class CreateMessageView(LoginRequiredMixin, View):
    """View for creating and sending system messages to other users."""
    
    template_name = 'users/create_message.html'
    
    def get(self, request):
        # Get all active users except the current user
        available_users = User.objects.filter(is_active=True)
        #available_users = User.objects.filter(is_active=True).exclude(id=request.user.id)
        
        return render(request, self.template_name, {
            'available_users': available_users
        })
    
    def post(self, request):
        try:
            # Get recipients from comma-separated string
            recipients_str = request.POST.get('recipients', '')
            if not recipients_str:
                messages.error(request, 'Please select at least one recipient.')
                return redirect('users:create-message')
            
            # Split the string into a list of IDs
            recipient_ids = [int(id_str) for id_str in recipients_str.split(',') if id_str.strip()]
            
            # Get message details
            title = request.POST.get('title')
            message_content = request.POST.get('message')
            priority = request.POST.get('priority', 'medium')
            
            # Create message for each recipient
            recipients = User.objects.filter(id__in=recipient_ids)
            for recipient in recipients:
                SystemMessage.create_message(
                    user=recipient,
                    title=title,
                    message=message_content,
                    priority=priority,
                    source_app='users',
                    source_model='User',
                    source_id=str(request.user.id)
                )
            
            messages.success(request, f'Message sent to {len(recipients)} recipient(s).')
            return redirect('users:messages')
            
        except Exception as e:
            messages.error(request, f'Error sending message: {str(e)}')
            return redirect('users:create-message')

@login_required
def user_settings_view(request):
    """View for displaying and managing user settings."""
    # Get all settings for the current user
    current_user_settings = UserSettings.get_all_settings(request.user)
    
    # Get all possible setting types
    setting_types = UserSetting.objects.values_list('setting_type', flat=True).distinct()
    
    context = {
        'settings': current_user_settings,
        'setting_types': list(setting_types),
    }
    
    return render(request, 'users/settings_view.html', context)


@login_required
def password_change_view(request):
    """View for changing user password"""
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, data=request.POST)
        if form.is_valid():
            new_password = form.cleaned_data['new_password1']
            request.user.set_password(new_password)
            request.user.save()
            messages.success(request, 'Your password has been changed successfully.')
            return redirect('users:profile')
    else:
        form = PasswordChangeForm(request.user)
    
    context = {
        'title': 'Change Password',
        'form': form,
    }
    
    return render(request, 'users/password_change.html', context)


@login_required
def password_set_view(request):
    """View for setting initial password for users without one"""
    # Check if user already has a password
    if request.user.has_usable_password():
        messages.info(request, 'You already have a password set.')
        return redirect('users:profile')
    
    if request.method == 'POST':
        form = PasswordSetForm(data=request.POST)
        if form.is_valid():
            new_password = form.cleaned_data['new_password1']
            request.user.set_password(new_password)
            request.user.save()
            messages.success(request, 'Your password has been set successfully.')
            return redirect('users:profile')
    else:
        form = PasswordSetForm()
    
    context = {
        'title': 'Set Password',
        'form': form,
    }
    
    return render(request, 'users/password_set.html', context)


def oauth_migration_view(request):
    """View for OAuth users to set passwords - email lookup step"""
    if request.method == 'POST':
        form = EmailLookupForm(data=request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            try:
                user = User.objects.get(email=email)
                # Store user ID in session for next step
                request.session['oauth_migration_user_id'] = user.id
                messages.success(request, f'Account found for {email}. Please set your password.')
                return redirect('users:oauth_password_set')
            except User.DoesNotExist:
                messages.error(request, f'No account found with email: {email}')
        else:
            # Form has validation errors
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = EmailLookupForm()
    
    # Debug: Show all users with emails for testing
    all_users = User.objects.filter(email__isnull=False).exclude(email='')
    debug_info = f"Available emails: {', '.join([u.email for u in all_users[:5]])}"  # Show first 5
    
    context = {
        'title': 'Set Your Password',
        'form': form,
        'debug_info': debug_info,
    }
    
    return render(request, 'users/oauth_migration.html', context)


def oauth_password_set_view(request):
    """View for OAuth users to set passwords - password entry step"""
    # Check if user ID is in session
    user_id = request.session.get('oauth_migration_user_id')
    if not user_id:
        messages.error(request, 'Please start the password setup process from the beginning.')
        return redirect('users:oauth_migration')
    
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        messages.error(request, 'Invalid user. Please start over.')
        return redirect('users:oauth_migration')
    
    if request.method == 'POST':
        form = OAuthPasswordSetForm(user, data=request.POST)
        if form.is_valid():
            new_password = form.cleaned_data['new_password1']
            user.set_password(new_password)
            user.save()
            # Clear session
            del request.session['oauth_migration_user_id']
            # Log the user in with the first available backend
            backend = settings.AUTHENTICATION_BACKENDS[0] if settings.AUTHENTICATION_BACKENDS else 'django.contrib.auth.backends.ModelBackend'
            auth_login(request, user, backend=backend)
            messages.success(request, f'Welcome back, {user.first_name or user.username}! Your password has been set successfully.')
            return redirect('index')
        else:
            # Form has validation errors
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = OAuthPasswordSetForm(user)
    
    # Show appropriate message based on whether user already has a password
    if user.has_usable_password():
        password_status = "update your password"
    else:
        password_status = "set your password"
    
    context = {
        'title': 'Set Your Password',
        'form': form,
        'user': user,
        'password_status': password_status,
    }
    
    return render(request, 'users/oauth_password_set.html', context)


# ---------------------------------------------------------------------------
# Portal dashboard APIs
# ---------------------------------------------------------------------------


@login_required
@require_http_methods(["GET"])
def portal_dashboard_data(request):
    """Return the aggregated portal context for async refreshes."""
    portal_context = build_portal_context(request.user)
    announcements = [
        {
            'id': announcement.id,
            'title': announcement.title,
            'content': announcement.content,
            'posted_at': announcement.posted_at.isoformat(),
            'posted_by': announcement.posted_by.get_full_name() or announcement.posted_by.username,
        }
        for announcement in Announcement.objects.select_related('posted_by').order_by('-posted_at')[:10]
    ]
    portal_context['announcements'] = announcements
    return JsonResponse(portal_context)


@login_required
@require_http_methods(["GET", "POST"])
def portal_sections_api(request):
    """List or create/update portal sections."""
    if request.method == "GET":
        sections = [serialize_section(section, request.user) for section in get_visible_sections(request.user)]
        return JsonResponse({'sections': sections})

    # Only superusers can create/update sections
    if not request.user.is_superuser:
        return JsonResponse({'error': 'Only superusers can modify sections.'}, status=403)

    data = _request_data(request)
    instance = None
    section_id = data.get('id') or data.get('section_id')
    if section_id:
        instance = get_object_or_404(PortalSection, pk=section_id)

    payload = data.copy() if isinstance(data, (dict, QueryDict)) else dict(data)
    if isinstance(payload, QueryDict):
        payload = payload.copy()

    editor_ids = _as_int_list(payload.get('editors'))
    if isinstance(payload, QueryDict):
        payload.setlist('editors', [str(pk) for pk in editor_ids])
    else:
        payload['editors'] = editor_ids

    form = PortalSectionForm(payload, instance=instance)
    if form.is_valid():
        section = form.save(commit=False)
        if not section.pk:
            section.created_by = request.user
        section.save()
        form.save_m2m()
        if request.user.is_authenticated:
            section.editors.add(request.user)
        return JsonResponse({'section': serialize_section(section, request.user)})
    return JsonResponse({'errors': form.errors}, status=400)


@login_required
@require_http_methods(["POST"])
def portal_section_delete(request, section_id):
    """Delete an existing portal section."""
    # Only superusers can delete sections
    if not request.user.is_superuser:
        return JsonResponse({'error': 'Only superusers can delete sections.'}, status=403)
    section = get_object_or_404(PortalSection, pk=section_id)
    section.delete()
    return JsonResponse({'success': True})


@login_required
@require_http_methods(["POST"])
def portal_resource_upsert(request):
    """Create or update a resource inside a portal section."""
    data = request.POST.copy() if request.POST else _request_data(request)
    files = request.FILES if request.FILES else None
    resource_id = data.get('id') or data.get('resource_id')
    section_id = data.get('section') or data.get('section_id')

    if not section_id and not resource_id:
        return JsonResponse({'error': 'Section is required to add resources.'}, status=400)

    instance = None
    if resource_id:
        instance = get_object_or_404(PortalResource, pk=resource_id)
        section = instance.section
    else:
        section = get_object_or_404(PortalSection, pk=section_id)

    if not _is_section_editor(request.user, section):
        return JsonResponse({'error': 'You do not have permission to manage this section.'}, status=403)

    # Enforce that only superusers can add/update file-type resources
    requested_type = (data.get('resource_type') or '').strip().lower()
    is_file_upload = (requested_type == 'file') or (bool(files) and ('file' in files))
    if instance and getattr(instance, 'resource_type', '') == 'file':
        is_file_upload = True
    if is_file_upload and not request.user.is_superuser:
        return JsonResponse({'error': 'Only superusers can add or modify file resources.'}, status=403)

    payload = data.copy() if isinstance(data, QueryDict) else dict(data)
    if isinstance(payload, QueryDict):
        payload = payload.copy()
        payload['section'] = str(section.id)
    else:
        payload['section'] = section.id

    form = PortalResourceForm(payload, files, instance=instance)
    if form.is_valid():
        resource = form.save(commit=False)
        resource.uploaded_by = request.user
        resource.save()
        form.save_m2m()
        return JsonResponse({'resource': serialize_resource(resource), 'section_id': section.id})
    return JsonResponse({'errors': form.errors}, status=400)


@login_required
@require_http_methods(["POST"])
def portal_resource_delete(request, resource_id):
    """Delete a portal resource."""
    resource = get_object_or_404(PortalResource, pk=resource_id)
    # File resources can be deleted by superusers only; others follow section editor rules
    if resource.resource_type == 'file':
        if not request.user.is_superuser:
            return JsonResponse({'error': 'Only superusers can delete file resources.'}, status=403)
    else:
        if not _is_section_editor(request.user, resource.section):
            return JsonResponse({'error': 'You do not have permission to delete this resource.'}, status=403)
    resource.delete()
    return JsonResponse({'success': True})


@login_required
@require_http_methods(["POST"])
def portal_task_create(request):
    """Create a task that can later be dropped onto the calendar."""
    data = _request_data(request)
    payload = data.copy() if isinstance(data, QueryDict) else dict(data)
    form = WorkCalendarTaskForm(payload)
    if form.is_valid():
        task = form.save(commit=False)
        task.owner = request.user
        task.save()
        return JsonResponse({'task': serialize_task(task)}, status=201)
    return JsonResponse({'errors': form.errors}, status=400)


@login_required
@require_http_methods(["POST"])
def portal_event_create(request):
    """Create a work calendar event from structured input."""
    data = request.POST.copy() if request.POST else _request_data(request)
    payload = data.copy() if isinstance(data, QueryDict) else dict(data)

    form = WorkCalendarEventForm(data=payload)
    if form.is_valid():
        event = form.save(commit=False)
        event.organizer = request.user
        if event.predicted_attendance is None:
            event.predicted_attendance = _calculate_predicted_attendance(request.user, event.kind)
        event.save()
        # Optional recurrence payload
        _upsert_recurrence_for_event(event, payload)
        return JsonResponse({'event': serialize_event(event, user=request.user)}, status=201)
    return JsonResponse({'errors': form.errors}, status=400)


@login_required
@require_http_methods(["GET"])
def portal_event_feed(request):
    """Retrieve events for the calendar widget.
    If 'start' and 'end' query params are provided, returns events overlapping that range.
    Otherwise, falls back to 'days' ahead from now (default 14).
    """
    start_param = request.GET.get('start')
    end_param = request.GET.get('end')

    if start_param and end_param:
        try:
            start_dt = _parse_datetime_string(start_param)
            end_dt = _parse_datetime_string(end_param)
        except Exception:
            return JsonResponse({'error': 'Invalid start or end datetime.'}, status=400)

        base_qs = WorkCalendarEvent.objects.select_related('organizer', 'section').prefetch_related('tasks', 'attendance_records', 'attachments')
        from django.db.models import Q
        # Company-wide visibility: show all non-private events, plus the user's own private events
        visibility = Q(is_private=False) | Q(organizer=request.user)
        qs = base_qs.filter(
            visibility,
            recurrence__isnull=True,  # non-recurring base events here
            end_at__gte=start_dt,
            start_at__lte=end_dt,
        ).distinct().order_by('start_at')
        # Recurring events -> expand occurrences
        rr = WorkCalendarEvent.objects.select_related('organizer', 'section', 'recurrence').prefetch_related('attachments').filter(
            visibility,
            recurrence__isnull=False,
        )
        events_payload = [serialize_event(ev, user=request.user) for ev in qs]
        for ev in rr:
            events_payload.extend(_expand_recurrences(ev, start_dt, end_dt, user=request.user))
        return JsonResponse({'events': events_payload})

    days = request.GET.get('days', 14)
    try:
        days = int(days)
    except (TypeError, ValueError):
        days = 14
    events = upcoming_events_for_user(request.user, days_ahead=days)
    return JsonResponse({'events': [serialize_event(event, user=request.user) for event in events]})


@login_required
@require_http_methods(["POST"])
def portal_event_delete(request, event_id):
    """Delete a calendar event (creator or superuser only)."""
    event = get_object_or_404(WorkCalendarEvent, pk=event_id)
    # Only the organizer can delete their own events (superusers cannot delete others)
    if event.organizer_id != request.user.id:
        return JsonResponse({'error': 'Permission denied.'}, status=403)
    event.delete()
    return JsonResponse({'success': True})


@login_required
@require_http_methods(["POST"])
def portal_event_update(request, event_id):
    """Update a calendar event — only the organizer can edit."""
    event = get_object_or_404(WorkCalendarEvent, pk=event_id)
    if event.organizer_id != request.user.id:
        return JsonResponse({'error': 'Permission denied.'}, status=403)

    data = _request_data(request)
    payload = data.copy() if isinstance(data, QueryDict) else dict(data)

    # Ensure organizer and immutable bits remain correct
    form = WorkCalendarEventForm(data=payload, instance=event)
    if form.is_valid():
        updated = form.save(commit=False)
        updated.organizer_id = event.organizer_id
        updated.save()
        _upsert_recurrence_for_event(updated, payload)
        return JsonResponse({'event': serialize_event(updated, user=request.user)})
    return JsonResponse({'errors': form.errors}, status=400)


def _upsert_recurrence_for_event(event, payload):
    """Create/update a simple recurrence rule for the event if provided.
    Supports freq 'weekly' (with byweekday list 0=Mon..6=Sun) and 'daily'.
    """
    freq = (payload.get('recurrence_freq') or '').strip().lower()
    if not freq or freq == 'none':
        # Remove existing rule if any
        try:
            if hasattr(event, 'recurrence') and event.recurrence:
                event.recurrence.delete()
        except RecurrenceRule.DoesNotExist:
            pass
        return
    interval = int(payload.get('recurrence_interval') or 1)
    byweekday = payload.get('recurrence_byweekday') or []
    if isinstance(byweekday, str):
        # Accept comma-separated values
        byweekday = [int(x) for x in byweekday.split(',') if x.strip().isdigit()]
    try:
        rule, _ = RecurrenceRule.objects.update_or_create(
            event=event,
            defaults={
                'freq': freq,
                'interval': max(1, interval),
                'byweekday': byweekday if isinstance(byweekday, (list, tuple)) else [],
            }
        )
    except Exception:
        # Be defensive; do not break event creation if recurrence invalid
        pass


def _expand_recurrences(event, range_start, range_end, user=None):
    """Generate occurrence payloads for a recurring event within the requested range."""
    rule = getattr(event, 'recurrence', None)
    if not rule:
        return []
    occurrences = []
    from datetime import datetime, timedelta
    start = event.start_at
    end = event.end_at
    # Normalize to start of the day for iteration
    cur_date = range_start.date()
    last_date = range_end.date()
    # Map Python weekday: Monday=0..Sunday=6; our byweekday uses same
    if rule.freq == 'weekly':
        # Iterate each day in range, include when weekday matches and step matches interval
        # Anchor interval against the event start week number
        from datetime import date
        def weeks_between(d1, d2):
            return int((d2 - d1).days // 7)
        anchor_monday = (start.date() - timedelta(days=start.weekday()))
        d = cur_date
        while d <= last_date:
            if rule.byweekday and d.weekday() in rule.byweekday:
                weeks = weeks_between(anchor_monday, d - timedelta(days=d.weekday()))
                if weeks % max(1, rule.interval) == 0:
                    # occurrence datetime spans
                    occ_start = timezone.make_aware(datetime.combine(d, start.timetz())) if timezone.is_naive(start) else start.replace(year=d.year, month=d.month, day=d.day)
                    duration = end - start
                    occ_end = occ_start + duration
                    if occ_end >= range_start and occ_start <= range_end:
                        ev = serialize_event(event, user=user)
                        ev['start'] = occ_start.isoformat()
                        ev['end'] = occ_end.isoformat()
                        occurrences.append(ev)
            d += timedelta(days=1)
    elif rule.freq == 'daily':
        d = cur_date
        step = max(1, rule.interval)
        delta_days = 0
        from datetime import datetime, timedelta
        while d <= last_date:
            if delta_days % step == 0:
                occ_start = timezone.make_aware(datetime.combine(d, start.timetz())) if timezone.is_naive(start) else start.replace(year=d.year, month=d.month, day=d.day)
                duration = end - start
                occ_end = occ_start + duration
                if occ_end >= range_start and occ_start <= range_end:
                    ev = serialize_event(event, user=user)
                    ev['start'] = occ_start.isoformat()
                    ev['end'] = occ_end.isoformat()
                    occurrences.append(ev)
            d += timedelta(days=1)
            delta_days += 1
    return occurrences


@login_required
@require_http_methods(["GET"])
def portal_event_detail(request, event_id):
    """Return serialized event with attachments (visibility-checked)."""
    ev = get_object_or_404(WorkCalendarEvent.objects.select_related('organizer').prefetch_related('attachments'), pk=event_id)
    if ev.is_private and ev.organizer_id != request.user.id:
        return JsonResponse({'error': 'Permission denied.'}, status=403)
    return JsonResponse({'event': serialize_event(ev, user=request.user)})


@login_required
@require_http_methods(["POST"])
def portal_event_attachment_upsert(request, event_id):
    """Create a new attachment (file or link) for an event (organizer only)."""
    event = get_object_or_404(WorkCalendarEvent, pk=event_id)
    if event.organizer_id != request.user.id:
        return JsonResponse({'error': 'Permission denied.'}, status=403)
    data = request.POST if request.method == 'POST' else _request_data(request)
    files = request.FILES if hasattr(request, 'FILES') else None
    form = EventAttachmentForm(data, files)
    if form.is_valid():
        att = form.save(commit=False)
        att.event = event
        att.uploaded_by = request.user
        att.save()
        return JsonResponse({'attachment': {
            'id': att.id,
            'title': att.title or (att.file.name.split('/')[-1] if att.file else att.link_url),
            'attachment_type': att.attachment_type,
            'url': att.get_absolute_url(),
        }})
    return JsonResponse({'errors': form.errors}, status=400)


@login_required
@require_http_methods(["POST"])
def portal_event_attachment_delete(request, attachment_id):
    att = get_object_or_404(EventAttachment.objects.select_related('event'), pk=attachment_id)
    if att.event.organizer_id != request.user.id:
        return JsonResponse({'error': 'Permission denied.'}, status=403)
    att.delete()
    return JsonResponse({'success': True})


@login_required
@require_http_methods(["POST"])
def portal_nlp_schedule(request):
    """Parse natural language scheduling requests and optionally auto-schedule."""
    data = _request_data(request)
    query = (data.get('query') or '').strip()
    if not query:
        return JsonResponse({'error': 'Describe what you want to schedule.'}, status=400)

    auto_schedule = data.get('auto_schedule', False)
    request_record = NaturalLanguageScheduleRequest.objects.create(user=request.user, raw_text=query)
    parse_result = _parse_natural_language_request(query, request.user)

    candidate_start, candidate_end, adjusted = _next_available_slot(
        request.user,
        parse_result['start'],
        parse_result['duration'],
        buffer_minutes=data.get('buffer_minutes', 15),
    )

    response_payload = {
        'query': query,
        'suggested_start': candidate_start.isoformat(),
        'suggested_end': candidate_end.isoformat(),
        'attendees': parse_result['attendees'],
        'duration': parse_result['duration'],
        'adjusted_for_conflicts': adjusted,
        'diagnostics': parse_result['diagnostics'],
    }

    request_record.status = 'parsed'
    request_record.interpreted_start = candidate_start
    request_record.interpreted_end = candidate_end
    request_record.duration_minutes = parse_result['duration']
    request_record.attendees = parse_result['attendees']
    request_record.diagnostics = parse_result['diagnostics']
    request_record.save(update_fields=['status', 'interpreted_start', 'interpreted_end', 'duration_minutes', 'attendees', 'diagnostics', 'updated_at'])

    if auto_schedule:
        title = data.get('title') or query[:120]
        description = data.get('description', '')
        kind = data.get('kind', 'meeting')
        priority = data.get('priority', 'normal')
        energy_required = data.get('energy_required', 'moderate')
        focus_block = data.get('focus_block', False)

        event = WorkCalendarEvent.objects.create(
            title=title,
            description=description,
            kind=kind,
            start_at=candidate_start,
            end_at=candidate_end,
            organizer=request.user,
            priority=priority,
            energy_required=energy_required,
            focus_block=focus_block,
            focus_reason=data.get('focus_reason', ''),
            requires_travel=data.get('requires_travel', False),
            created_via_nlp=True,
            metadata={
                'nlp_query': query,
                'attendee_names': parse_result['attendees'],
                'adjusted_for_conflicts': adjusted,
            },
        )
        event.predicted_attendance = _calculate_predicted_attendance(request.user, kind)
        event.save()
        request_record.mark_scheduled(event, diagnostics=parse_result['diagnostics'])
        microbreak = _auto_insert_microbreak(request.user, event)

        response_payload['event'] = serialize_event(event, user=request.user)
        response_payload['microbreak'] = serialize_microbreak(microbreak) if microbreak else None

    return JsonResponse(response_payload, status=201 if auto_schedule else 200)


@login_required
@require_http_methods(["POST"])
def portal_microbreak_create(request):
    """Allow users to manually insert a micro-break."""
    data = _request_data(request)
    try:
        start = _parse_datetime_string(data.get('start'))
        end = _parse_datetime_string(data.get('end'))
    except Exception:
        return JsonResponse({'error': 'Invalid start or end time for micro-break.'}, status=400)

    if end <= start:
        return JsonResponse({'error': 'Micro-break end must be after start.'}, status=400)

    label = data.get('label') or 'Micro-break'
    related_event_id = data.get('related_event_id')
    related_event = None
    if related_event_id:
        related_event = get_object_or_404(WorkCalendarEvent, pk=related_event_id)

    microbreak = ScheduledMicroBreak.objects.create(
        user=request.user,
        start_at=start,
        end_at=end,
        label=label,
        insertion_mode='manual',
        related_event=related_event,
        notes=data.get('notes', ''),
    )
    return JsonResponse({'microbreak': serialize_microbreak(microbreak)}, status=201)


@login_required
@require_http_methods(["GET"])
def portal_microbreak_feed(request):
    """Return upcoming micro-breaks for the current user."""
    upcoming = upcoming_microbreaks(request.user)
    return JsonResponse({'microbreaks': [serialize_microbreak(break_obj) for break_obj in upcoming]})


def custom_password_reset(request):
    """Custom password reset that redirects OAuth users to migration flow"""
    if request.method == 'POST':
        email = request.POST.get('email')
        if email:
            try:
                user = User.objects.get(email=email)
                if not user.has_usable_password():
                    # OAuth user - redirect to migration flow
                    request.session['oauth_migration_user_id'] = user.id
                    messages.info(request, 'OAuth user detected. Redirecting to password setup...')
                    return redirect('users:oauth_password_set')
                else:
                    # Regular user - use standard password reset
                    return redirect('users:password_reset')
            except User.DoesNotExist:
                messages.error(request, 'No account found with this email address.')
    
    # Show email form
    form = EmailLookupForm()
    context = {
        'title': 'Password Reset',
        'form': form,
    }
    return render(request, 'users/custom_password_reset.html', context)
def _import_sharepoint_xlsx_core(workbook, organizer):
    from django.utils import timezone as dj_tz
    ws = workbook.active
    # Map headers
    headers = {}
    for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)), start=1):
        key = (str(cell) if cell is not None else '').strip().lower()
        headers[key] = idx
    def col(name):
        return headers.get(name.lower())
    required = ['start date', 'end date', 'title']
    if not all(h.lower() in headers for h in required):
        raise ValueError('Missing required headers. Need: Start Date, End Date, Title')

    created = 0
    errors = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            raw_title = row[col('title')-1] if col('title') else ''
            title = (raw_title or '').strip()
            if not title:
                continue
            s = row[col('start date')-1] if col('start date') else None
            e = row[col('end date')-1] if col('end date') else None
            # Optional fields
            all_day_col = col('all day event') or col('all day')
            all_day_val = (row[all_day_col-1] if all_day_col else None)
            category_col = col('category') or col('categories')
            category_val = (row[category_col-1] if category_col else None)

            from datetime import datetime, timedelta, time, date
            if s is None:
                continue
            if isinstance(s, date) and not isinstance(s, datetime):
                s = datetime.combine(s, time(0,0,0))
            # Detect all-day: either explicit or date-only values
            is_all_day = False
            if all_day_val is not None:
                is_all_day = str(all_day_val).strip().lower() in ('1','true','yes','y')
            if e is None:
                if is_all_day:
                    e = s.replace(hour=23, minute=59, second=0)
                else:
                    e = s + timedelta(hours=1)
            elif isinstance(e, date) and not isinstance(e, datetime):
                # SharePoint exports often give date-only for all day
                e = datetime.combine(e, time(23,59,0))
                is_all_day = True

            if dj_tz.is_naive(s):
                s = dj_tz.make_aware(s, dj_tz.get_current_timezone())
            if dj_tz.is_naive(e):
                e = dj_tz.make_aware(e, dj_tz.get_current_timezone())
            if e <= s:
                e = s + timedelta(minutes=30)

            # Map category to kind
            kind = 'meeting'
            cat = (str(category_val).lower() if category_val else '')
            if any(k in cat for k in ['holiday', 'wfh', 'work from home', 'vacation', 'sick', 'personal']):
                kind = 'personal'
            elif 'training' in cat:
                kind = 'training'
            elif 'travel' in cat:
                kind = 'travel'
            elif 'focus' in cat:
                kind = 'focus'

            ev = WorkCalendarEvent(
                title=title,
                description='',
                kind=kind,
                start_at=s,
                end_at=e,
                organizer=organizer,
                priority='normal',
                is_private=False,
            )
            ev.full_clean()
            ev.save()
            created += 1
        except Exception:
            errors += 1
    return created, errors


@login_required
@require_http_methods(["POST"])
def portal_import_sharepoint_xlsx(request):
    """Import SharePoint calendar export (XLSX) via API. Returns JSON counts."""
    from openpyxl import load_workbook
    from io import BytesIO
    import os

    f = request.FILES.get('file') if hasattr(request, 'FILES') else None
    try:
        if f:
            wb = load_workbook(filename=BytesIO(f.read()), data_only=True)
        else:
            path = os.path.join(os.path.dirname(__file__), 'AllItems.xlsx')
            if not os.path.exists(path):
                return JsonResponse({'error': 'No file uploaded and default AllItems.xlsx not found.'}, status=400)
            wb = load_workbook(filename=path, data_only=True)
        created, errors = _import_sharepoint_xlsx_core(wb, request.user)
        return JsonResponse({'imported': created, 'skipped': errors})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


def _is_staff(user):
    return bool(user and user.is_authenticated and (user.is_staff or user.is_superuser))


@login_required
@user_passes_test(_is_staff)
@require_http_methods(["GET", "POST"])
def sharepoint_import_ui(request):
    """Simple staff UI to upload a SharePoint XLSX and import events."""
    context = {'result': None, 'error': None, 'title': 'Import Calendar (SharePoint XLSX)'}
    if request.method == 'POST':
        try:
            from openpyxl import load_workbook
            from io import BytesIO
            f = request.FILES.get('file')
            if not f:
                context['error'] = 'Please choose an .xlsx file.'
            else:
                wb = load_workbook(filename=BytesIO(f.read()), data_only=True)
                created, errors = _import_sharepoint_xlsx_core(wb, request.user)
                context['result'] = {'imported': created, 'skipped': errors}
        except Exception as e:
            context['error'] = str(e)
    return render(request, 'users/import_sharepoint.html', context)


@login_required
@require_http_methods(["GET"])
def portal_events_export_csv(request):
    """Export events as CSV for the visible range or all if not provided.
    Exposes only non-private events plus the user's own private events.
    """
    import csv
    from django.http import HttpResponse
    start_param = request.GET.get('start')
    end_param = request.GET.get('end')
    from django.db.models import Q
    qs = WorkCalendarEvent.objects.select_related('organizer').all()
    visibility = Q(is_private=False) | Q(organizer=request.user)
    qs = qs.filter(visibility)
    if start_param and end_param:
        try:
            start_dt = _parse_datetime_string(start_param)
            end_dt = _parse_datetime_string(end_param)
            qs = qs.filter(end_at__gte=start_dt, start_at__lte=end_dt)
        except Exception:
            pass
    qs = qs.order_by('start_at')

    resp = HttpResponse(content_type='text/csv')
    resp['Content-Disposition'] = 'attachment; filename="events_export.csv"'
    writer = csv.writer(resp)
    writer.writerow(['id','title','start','end','all_day','kind','location','is_private','organizer'])
    for ev in qs:
        all_day = (ev.start_at.hour == 0 and ev.end_at.hour == 23 and ev.end_at.minute >= 59)
        writer.writerow([
            ev.id,
            ev.title,
            ev.start_at.isoformat(),
            ev.end_at.isoformat(),
            'true' if all_day else 'false',
            ev.kind,
            ev.location,
            'true' if ev.is_private else 'false',
            ev.organizer.get_full_name() or ev.organizer.username,
        ])
    return resp
