"""Shared Contract Log (masterlog) export generation.

Used by both the synchronous XLSX export view and the async
ContractLogExportJob worker (contracts/management/commands/process_contract_log_exports.py)
so the two paths can never drift out of sync.
"""
from decimal import Decimal
from io import BytesIO

from django.db.models import Subquery, OuterRef, Sum, Count

from ..models import Clin, ClinSplit


def build_export_queryset(company, filters_dict):
    """Return the materialized, filtered, annotated Clin list for a Contract Log export."""
    # Imported here (not at module load) to avoid a circular import: contract_log_views
    # imports generate_contract_log_xlsx from this module for the synchronous export view.
    from ..views.contract_log_views import _apply_log_filters

    clins = Clin.objects.filter(company=company).select_related(
        'contract',
        'contract__buyer',
        'contract__contract_type',
        'contract__status',
        'contract__idiq_contract',
        'contract__special_payment_terms',
        'supplier',
        'nsn'
    ).prefetch_related(
        'clinacknowledgment_set',
    ).annotate(
        ppi_split_paid=Subquery(
            ClinSplit.objects
            .filter(clin__contract_id=OuterRef('contract_id'), company_name__iexact='PPI')
            .order_by()
            .values('clin__contract_id')
            .annotate(total=Sum('split_paid'))
            .values('total')[:1]
        ),
        statz_split_paid=Subquery(
            ClinSplit.objects
            .filter(clin__contract_id=OuterRef('contract_id'), company_name__iexact='STATZ')
            .order_by()
            .values('clin__contract_id')
            .annotate(total=Sum('split_paid'))
            .values('total')[:1]
        ),
        notes_count=Count('notes', distinct=True),
    ).order_by('contract__award_date', 'contract__po_number', 'item_number')

    clins = _apply_log_filters(clins, filters_dict, company)
    return list(clins)


def generate_contract_log_xlsx(company, filters_dict):
    """Build the Contract Log XLSX workbook and return (file_bytes, row_count)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    clins = build_export_queryset(company, filters_dict)

    wb = Workbook()
    ws = wb.active
    ws.title = 'MASTER CONTRACT LOG Export'

    bold = Font(bold=True)
    header_fill = PatternFill('solid', fgColor='F2F2F2')
    center = Alignment(horizontal='center', vertical='center')
    thin = Side(border_style='thin', color='DDDDDD')
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    company_name = getattr(company, 'name', 'STATZ Corporation')
    ws.append([company_name])
    now = datetime.now()
    ws.append([f"Government Contracting Log - Master List", '', '', f"Export @ {now.strftime('%I:%M:%S %p')}"])
    ws.append([])

    headers = [
        'Open', 'PO #', 'IDIQ Contract #', 'Contract', 'Buyer', 'Type', 'CLIN #',
        'Supplier', 'Cage Code', 'Award Date', 'Contract Status', 'NSN', 'Item Description',
        'I&A', 'PO to Sub', 'Sub Reply', 'PO to QAR', 'FOB', 'QDD', 'CDD', 'Qty / UOM',
        'Ship Date', 'Ship Qty', 'Sub PO $', 'Sub Paid $', 'Item Value', 'Terms', 'Contract $',
        'Customer Payment $', 'Date Pay Recv', 'Plan Gross $', 'Actual Paid PPI $', 'Actual STATZ $',
        'Notes'
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=4, column=col)
        c.font = bold
        c.fill = header_fill
        c.alignment = center
        c.border = border

    seen_contracts = set()
    # 1-indexed columns for currency formatting:
    # Sub PO$(24), Sub Paid$(25), Item Value(26), Contract$(28), Customer Pay$(29),
    # Plan Gross$(31), PPI$(32), STATZ$(33)
    money_cols = {24, 25, 26, 28, 29, 31, 32, 33}

    for clin in clins:
        ack = clin.clinacknowledgment_set.first()
        first_for_contract = clin.contract_id not in seen_contracts
        if first_for_contract:
            seen_contracts.add(clin.contract_id)

        ppi_split_paid = Decimal('0')
        statz_split_paid = Decimal('0')
        if clin.contract_id and first_for_contract:
            ppi_split_paid = clin.ppi_split_paid or Decimal('0')
            statz_split_paid = clin.statz_split_paid or Decimal('0')

        if clin.contract and clin.contract.status and getattr(clin.contract.status, 'description', '') == 'Canceled':
            status_char = 'X'
        elif clin.contract and clin.contract.date_closed:
            status_char = 'C'
        else:
            status_char = 'O'

        parts = []
        if clin.contract and getattr(clin.contract, 'status', None) and getattr(clin.contract.status, 'description', ''):
            parts.append(clin.contract.status.description)
        if not (ack and ack.po_to_supplier_bool):
            parts.append('PO NOT SENT YET;')
        if not (ack and ack.clin_reply_bool):
            parts.append('SUB REPLY NEEDED;')
        if not (ack and ack.po_to_qar_bool):
            parts.append('PO TO QAR NEEDED;')
        status_text = ' '.join(parts).strip()

        terms = ''
        if clin.contract and getattr(clin.contract, 'special_payment_terms', None):
            spt = clin.contract.special_payment_terms
            terms = getattr(spt, 'terms', None) or getattr(spt, 'code', '') or ''

        qty_uom = f"{float(clin.order_qty):g} {clin.uom or 'ea'}" if clin.order_qty not in (None, '') else ''

        row = [
            status_char,
            (clin.po_number or clin.clin_po_num or (clin.contract.po_number if clin.contract else '')),
            (clin.contract.idiq_contract.contract_number if clin.contract and clin.contract.idiq_contract else ''),
            clin.contract.contract_number if clin.contract else '',
            clin.contract.buyer.description if clin.contract and clin.contract.buyer else '',
            clin.contract.contract_type.description if clin.contract and clin.contract.contract_type else '',
            clin.item_number or '',
            clin.supplier.name if clin.supplier else '',
            clin.supplier.cage_code if clin.supplier else '',
            clin.contract.award_date.strftime('%m/%d/%Y') if clin.contract and clin.contract.award_date else '',
            status_text,
            clin.nsn.nsn_code if clin.nsn else '',
            clin.nsn.description if clin.nsn else '',
            clin.ia or '',
            1 if (ack and ack.po_to_supplier_bool) else 0,
            1 if (ack and ack.clin_reply_bool) else 0,
            1 if (ack and ack.po_to_qar_bool) else 0,
            clin.fob or '',
            clin.supplier_due_date.strftime('%m/%d/%Y') if clin.supplier_due_date else '',
            clin.due_date.strftime('%m/%d/%Y') if clin.due_date else '',
            qty_uom,
            clin.ship_date.strftime('%m/%d/%Y') if clin.ship_date else '',
            float(clin.ship_qty) if clin.ship_qty not in (None, '') else '',
            float(clin.quote_value) if clin.quote_value else '',
            float(clin.paid_amount) if clin.paid_amount else '',
            float(clin.item_value) if clin.item_value else '',
            terms,
            float(clin.contract.contract_value) if (first_for_contract and clin.contract and clin.contract.contract_value) else 0.0,
            float(clin.wawf_payment) if clin.wawf_payment else '',
            clin.wawf_recieved.strftime('%m/%d/%Y') if clin.wawf_recieved else '',
            float(clin.contract.plan_gross) if (first_for_contract and clin.contract and clin.contract.plan_gross is not None) else 0.0,
            float(ppi_split_paid) if (first_for_contract and ppi_split_paid) else 0.0,
            float(statz_split_paid) if (first_for_contract and statz_split_paid) else 0.0,
            int(clin.notes_count)
        ]
        ws.append(row)
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = border

    widths = [6, 10, 18, 18, 14, 12, 8, 24, 10, 12, 26, 12, 24, 8, 10, 10, 10, 8, 10, 10, 10, 10, 10, 12, 12, 12, 12, 14, 14, 14, 14, 14, 14, 30]
    for i in range(1, len(headers) + 1):
        try:
            w = widths[i - 1] if i - 1 < len(widths) else 12
            ws.column_dimensions[get_column_letter(i)].width = w
        except Exception:
            pass

    currency_fmt = '[$$-409]#,##0.00'
    for row in ws.iter_rows(min_row=5, min_col=1, max_col=len(headers), max_row=ws.max_row):
        for idx, cell in enumerate(row, start=1):
            if idx in money_cols and isinstance(cell.value, (int, float)):
                cell.number_format = currency_fmt

    ws.freeze_panes = 'A5'

    buff = BytesIO()
    wb.save(buff)
    buff.seek(0)

    return buff.getvalue(), len(clins)
