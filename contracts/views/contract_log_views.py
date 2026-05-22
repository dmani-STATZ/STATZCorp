from django.shortcuts import render, redirect
from django.views.generic import ListView
from django.utils.decorators import method_decorator
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, F, Value, CharField, Case, When, BooleanField, Exists, OuterRef, Subquery, Max, Count, Sum
from django.db.models.functions import Concat
from django.utils import timezone
import csv
import os
import subprocess
from django.conf import settings
import sys
from decimal import Decimal
import time
from datetime import date as _date_cls

from STATZWeb.decorators import conditional_login_required
from ..models import Clin, ClinAcknowledgment, ClinSplit, Supplier, ExportTiming, Buyer, ContractType


def _apply_log_filters(clins, params, active_company):
    """Apply all master-log GET filter params to a Clin queryset."""
    def _d(val):
        try:
            return _date_cls.fromisoformat(val.strip())
        except (ValueError, AttributeError):
            return None

    # ── existing filters ──────────────────────────────────────────────
    search = params.get('search', '').strip()
    if search:
        clins = clins.filter(
            Q(contract__contract_number__icontains=search) |
            Q(contract__po_number__icontains=search) |
            Q(contract__tab_num__icontains=search) |
            Q(supplier__name__icontains=search) |
            Q(nsn__nsn_code__icontains=search) |
            Q(item_number__icontains=search)
        ).distinct()

    status = params.get('status', '').strip()
    if status == 'open':
        clins = clins.filter(
            Q(contract__status__description='Open') &
            Q(contract__date_closed__isnull=True)
        )
    elif status == 'closed':
        clins = clins.filter(
            Q(contract__status__description='Closed') &
            Q(contract__date_closed__isnull=False)
        )
    elif status in ('canceled', 'cancelled'):
        clins = clins.filter(contract__status__description='Canceled')

    supplier = params.get('supplier', '').strip()
    if supplier:
        contract_ids = Clin.objects.filter(
            supplier_id=supplier,
            company=active_company,
        ).values_list('contract_id', flat=True).distinct()
        clins = clins.filter(contract_id__in=contract_ids)

    # ── new column filters ────────────────────────────────────────────
    po = params.get('po', '').strip()
    if po:
        clins = clins.filter(contract__po_number__icontains=po)

    idiq = params.get('idiq', '').strip()
    if idiq:
        clins = clins.filter(contract__idiq_contract__contract_number__icontains=idiq)

    contract_num = params.get('contract', '').strip()
    if contract_num:
        clins = clins.filter(contract__contract_number__icontains=contract_num)

    buyer = params.get('buyer', '').strip()
    if buyer:
        clins = clins.filter(contract__buyer_id=buyer)

    ctype = params.get('ctype', '').strip()
    if ctype:
        clins = clins.filter(contract__contract_type_id=ctype)

    clin_num = params.get('clin', '').strip()
    if clin_num:
        clins = clins.filter(item_number__icontains=clin_num)

    cage = params.get('cage', '').strip()
    if cage:
        clins = clins.filter(supplier__cage_code__icontains=cage)

    nsn = params.get('nsn', '').strip()
    if nsn:
        clins = clins.filter(nsn__nsn_code__icontains=nsn)

    desc = params.get('desc', '').strip()
    if desc:
        clins = clins.filter(nsn__description__icontains=desc)

    ia = params.get('ia', '').strip()
    if ia:
        clins = clins.filter(ia=ia)

    fob = params.get('fob', '').strip()
    if fob:
        clins = clins.filter(fob=fob)

    # Date ranges
    award_from = _d(params.get('award_from', ''))
    award_to   = _d(params.get('award_to', ''))
    if award_from:
        clins = clins.filter(contract__award_date__gte=award_from)
    if award_to:
        clins = clins.filter(contract__award_date__lte=award_to)

    qdd_from = _d(params.get('qdd_from', ''))
    qdd_to   = _d(params.get('qdd_to', ''))
    if qdd_from:
        clins = clins.filter(supplier_due_date__gte=qdd_from)
    if qdd_to:
        clins = clins.filter(supplier_due_date__lte=qdd_to)

    due_from = _d(params.get('due_from', ''))
    due_to   = _d(params.get('due_to', ''))
    if due_from:
        clins = clins.filter(due_date__gte=due_from)
    if due_to:
        clins = clins.filter(due_date__lte=due_to)

    ship_from = _d(params.get('ship_from', ''))
    ship_to   = _d(params.get('ship_to', ''))
    if ship_from:
        clins = clins.filter(ship_date__gte=ship_from)
    if ship_to:
        clins = clins.filter(ship_date__lte=ship_to)

    return clins


class ContractLogView(ListView):
    model = Clin
    template_name = 'contracts/contract_log_view.html'
    context_object_name = 'clins'
    paginate_by = 25
    allowed_per_page = (25, 50, 100)

    def get_paginate_by(self, queryset):
        """Allow per_page from GET (25, 50, 100) for easier scanning."""
        try:
            per = int(self.request.GET.get('per_page', self.paginate_by))
            if per in self.allowed_per_page:
                return per
        except (TypeError, ValueError):
            pass
        return self.paginate_by

    def get_queryset(self):
        clins = Clin.objects.filter(company=self.request.active_company).select_related(
            'contract',
            'contract__buyer',
            'contract__contract_type',
            'contract__status',
            'contract__idiq_contract',
            'supplier',
            'nsn'
        ).prefetch_related(
            'clinacknowledgment_set',
        ).annotate(
            ppi_split_paid=Subquery(
                ClinSplit.objects
                .filter(clin__contract_id=OuterRef('contract_id'), company_name__iexact='PPI')
                .order_by()
                .values('clin__contract_id')
                .annotate(total=Sum('split_paid'))
                .values('total')[:1]
            ),
            statz_split_paid=Subquery(
                ClinSplit.objects
                .filter(clin__contract_id=OuterRef('contract_id'), company_name__iexact='STATZ')
                .order_by()
                .values('clin__contract_id')
                .annotate(total=Sum('split_paid'))
                .values('total')[:1]
            ),
            # Flag first CLIN per contract via subquery (SQLite/SQL Server safe)
            is_first_for_contract=Case(
                When(
                    id=Subquery(
                        Clin.objects
                        .filter(contract_id=OuterRef('contract_id'))
                        .order_by('item_number', 'id')
                        .values('id')[:1]
                    ),
                    then=Value(True)
                ),
                default=Value(False),
                output_field=BooleanField(),
            ),
        ).order_by('contract__award_date', 'contract__po_number', 'item_number')

        return _apply_log_filters(clins, self.request.GET, self.request.active_company)

    def get(self, request, *args, **kwargs):
        # If no page is specified, calculate and redirect to the last page
        if 'page' not in request.GET:
            queryset = self.get_queryset()
            paginator = self.get_paginator(queryset, self.paginate_by)
            params = request.GET.copy()
            params['page'] = paginator.num_pages
            return redirect(f"{request.path}?{params.urlencode()}")
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Existing filter params
        context['search_query'] = self.request.GET.get('search', '')
        context['status_filter'] = self.request.GET.get('status', '')
        context['supplier_filter'] = self.request.GET.get('supplier', '')
        paginator = context.get('paginator')
        context['current_per_page'] = paginator.per_page if paginator else self.paginate_by
        context['allowed_per_page'] = self.allowed_per_page
        params = self.request.GET.copy()
        params.pop('page', None)
        context['pagination_query_string'] = params.urlencode()

        # Dropdown data for filter sidebar
        context['suppliers'] = Supplier.objects.all().order_by('name')
        context['buyers'] = Buyer.objects.all().order_by('description')
        context['contract_types'] = ContractType.objects.all().order_by('description')

        # Pass all new filter values back for sidebar pre-population
        for p in ['po', 'idiq', 'contract', 'buyer', 'ctype', 'clin', 'cage', 'nsn', 'desc',
                  'ia', 'fob', 'award_from', 'award_to', 'qdd_from', 'qdd_to',
                  'due_from', 'due_to', 'ship_from', 'ship_to']:
            context[f'filter_{p}'] = self.request.GET.get(p, '')

        # Build derived status text for each CLIN (used by UI Contract Status column)
        page_obj = context.get('page_obj')
        if page_obj:
            def build_status_text(clin):
                parts = []
                if getattr(clin.contract, 'status', None) and getattr(clin.contract.status, 'description', ''):
                    parts.append(clin.contract.status.description)
                ack = None
                try:
                    ack = clin.clinacknowledgment_set.first()
                except Exception:
                    ack = None
                if not (ack and ack.po_to_supplier_bool):
                    parts.append('PO NOT SENT YET;')
                if not (ack and ack.clin_reply_bool):
                    parts.append('SUB REPLY NEEDED;')
                if not (ack and ack.po_to_qar_bool):
                    parts.append('PO TO QAR NEEDED;')
                return ' '.join(parts).strip()

            for clin in page_obj.object_list:
                try:
                    clin.contract_status_text = build_status_text(clin)
                except Exception:
                    clin.contract_status_text = getattr(clin.contract.status, 'description', '') if getattr(clin, 'contract', None) and getattr(clin.contract, 'status', None) else ''

        if self.paginate_by is not None:
            paginator = context.get('paginator')
            page_obj = context.get('page_obj')
            if paginator and page_obj:
                context['is_paginated'] = True
                context['page_obj'] = page_obj
                context['paginator'] = paginator

        return context


@conditional_login_required
def export_contract_log(request):
    """Export contract log to CSV with all relevant fields."""
    start_time = time.time()

    clins = Clin.objects.filter(company=request.active_company).select_related(
        'contract',
        'contract__buyer',
        'contract__contract_type',
        'contract__status',
        'contract__idiq_contract',
        'supplier',
        'nsn'
    ).prefetch_related(
        'clinacknowledgment_set',
    ).order_by('contract__award_date', 'contract__po_number', 'item_number')

    clins = _apply_log_filters(clins, request.GET, request.active_company)
    filters_applied = {k: v for k, v in request.GET.items() if k not in ('page', 'per_page') and v}

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="contract_log.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'Open', 'PO #', 'IDIQ Contract #', 'Contract', 'Buyer', 'Type', 'CLIN #',
        'Supplier', 'Cage Code', 'Award Date', 'Contract Status', 'NSN', 'Item Description',
        'I&A', 'PO to Sub', 'Sub Reply', 'PO to QAR', 'FOB', 'QDD', 'CDD', 'Qty / UOM',
        'Ship Date', 'Ship Qty', 'Sub PO $', 'Sub Paid $', 'Item Value', 'Terms', 'Contract $',
        'Customer Payment $', 'Date Pay Recv', 'Plan Gross $', 'Actual Paid PPI $', 'Actual STATZ $',
        'Notes'
    ])

    total_rows = clins.count()
    seen_contracts = set()

    for clin in clins:
        acknowledgment = clin.clinacknowledgment_set.first()
        first_for_contract = clin.contract_id not in seen_contracts
        if first_for_contract:
            seen_contracts.add(clin.contract_id)

        if clin.contract and clin.contract.status and getattr(clin.contract.status, 'description', '') == 'Canceled':
            first_col_status = 'X'
        elif clin.contract and clin.contract.date_closed:
            first_col_status = 'C'
        else:
            first_col_status = 'O'

        po_num = clin.po_number or clin.clin_po_num or (clin.contract.po_number if clin.contract else '')

        terms_text = ''
        c_terms = getattr(clin.contract, 'special_payment_terms', None) if getattr(clin, 'contract', None) else None
        if c_terms is not None:
            terms_text = getattr(c_terms, 'terms', None) or getattr(c_terms, 'code', '') or ''

        status_parts = []
        if clin.contract and getattr(clin.contract, 'status', None) and getattr(clin.contract.status, 'description', ''):
            status_parts.append(clin.contract.status.description)
        if not (acknowledgment and acknowledgment.po_to_supplier_bool):
            status_parts.append('PO NOT SENT YET;')
        if not (acknowledgment and acknowledgment.clin_reply_bool):
            status_parts.append('SUB REPLY NEEDED;')
        if not (acknowledgment and acknowledgment.po_to_qar_bool):
            status_parts.append('PO TO QAR NEEDED;')
        contract_status_text = ' '.join(status_parts).strip()

        ppi_split_paid = Decimal('0')
        statz_split_paid = Decimal('0')
        if clin.contract_id and first_for_contract:
            ppi_split_paid = (
                ClinSplit.objects.filter(
                    clin__contract_id=clin.contract_id,
                    company_name__iexact='PPI',
                ).aggregate(t=Sum('split_paid'))['t'] or Decimal('0')
            )
            statz_split_paid = (
                ClinSplit.objects.filter(
                    clin__contract_id=clin.contract_id,
                    company_name__iexact='STATZ',
                ).aggregate(t=Sum('split_paid'))['t'] or Decimal('0')
            )

        qty_uom = f"{clin.order_qty:g} {clin.uom or 'ea'}" if clin.order_qty not in (None, '') else ''

        writer.writerow([
            first_col_status,
            po_num,
            (clin.contract.idiq_contract.contract_number if clin.contract and clin.contract.idiq_contract else ''),
            clin.contract.contract_number if clin.contract else '',
            clin.contract.buyer.description if clin.contract and clin.contract.buyer else '',
            clin.contract.contract_type.description if clin.contract and clin.contract.contract_type else '',
            clin.item_number or '',
            clin.supplier.name if clin.supplier else '',
            clin.supplier.cage_code if clin.supplier else '',
            clin.contract.award_date.strftime('%m/%d/%Y') if clin.contract and clin.contract.award_date else '',
            contract_status_text,
            clin.nsn.nsn_code if clin.nsn else '',
            clin.nsn.description if clin.nsn else '',
            clin.ia or '',
            '1' if acknowledgment and acknowledgment.po_to_supplier_bool else '0',
            '1' if acknowledgment and acknowledgment.clin_reply_bool else '0',
            '1' if acknowledgment and acknowledgment.po_to_qar_bool else '0',
            clin.fob or '',
            clin.supplier_due_date.strftime('%m/%d/%Y') if clin.supplier_due_date else '',
            clin.due_date.strftime('%m/%d/%Y') if clin.due_date else '',
            qty_uom,
            clin.ship_date.strftime('%m/%d/%Y') if clin.ship_date else '',
            f"{clin.ship_qty:g}" if clin.ship_qty not in (None, '') else '',
            f"${clin.quote_value:,.2f}" if clin.quote_value else '',
            f"${clin.paid_amount:,.2f}" if clin.paid_amount else '',
            f"${clin.item_value:,.2f}" if clin.item_value else '',
            terms_text,
            f"${clin.contract.contract_value:,.2f}" if (first_for_contract and clin.contract and clin.contract.contract_value) else '$0.00',
            f"${clin.wawf_payment:,.2f}" if clin.wawf_payment else '',
            clin.wawf_recieved.strftime('%m/%d/%Y') if clin.wawf_recieved else '',
            f"${clin.contract.plan_gross:,.2f}" if (first_for_contract and clin.contract and clin.contract.plan_gross is not None) else '$0.00',
            f"${ppi_split_paid:,.2f}" if (first_for_contract and ppi_split_paid) else '$0.00',
            f"${statz_split_paid:,.2f}" if (first_for_contract and statz_split_paid) else '$0.00',
            clin.notes.count()
        ])

    end_time = time.time()
    ExportTiming.objects.create(
        row_count=total_rows,
        export_time=(end_time - start_time),
        filters_applied=filters_applied
    )

    return response


@conditional_login_required
def export_contract_log_xlsx(request):
    """Export contract log to XLSX with workbook-like structure."""
    start_time = time.time()
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
    except Exception as e:
        return HttpResponse(
            f"Missing dependency for XLSX export: {e}. Please install 'openpyxl'.",
            status=500,
            content_type='text/plain'
        )

    clins = Clin.objects.filter(company=request.active_company).select_related(
        'contract',
        'contract__buyer',
        'contract__contract_type',
        'contract__status',
        'contract__idiq_contract',
        'contract__special_payment_terms',
        'supplier',
        'nsn'
    ).prefetch_related(
        'clinacknowledgment_set',
    ).order_by('contract__award_date', 'contract__po_number', 'item_number')

    clins = _apply_log_filters(clins, request.GET, request.active_company)
    filters_applied = {k: v for k, v in request.GET.items() if k not in ('page', 'per_page') and v}

    wb = Workbook()
    ws = wb.active
    ws.title = 'MASTER CONTRACT LOG Export'

    bold = Font(bold=True)
    header_fill = PatternFill('solid', fgColor='F2F2F2')
    center = Alignment(horizontal='center', vertical='center')
    thin = Side(border_style='thin', color='DDDDDD')
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    company_name = getattr(getattr(request, 'active_company', None), 'name', 'STATZ Corporation')
    ws.append([company_name])
    from datetime import datetime
    now = datetime.now()
    ws.append([f"Government Contracting Log - Master List", '', '', f"Export @ {now.strftime('%I:%M:%S %p')}"])
    ws.append([])

    headers = [
        'Open', 'PO #', 'IDIQ Contract #', 'Contract', 'Buyer', 'Type', 'CLIN #',
        'Supplier', 'Cage Code', 'Award Date', 'Contract Status', 'NSN', 'Item Description',
        'I&A', 'PO to Sub', 'Sub Reply', 'PO to QAR', 'FOB', 'QDD', 'CDD', 'Qty / UOM',
        'Ship Date', 'Ship Qty', 'Sub PO $', 'Sub Paid $', 'Item Value', 'Terms', 'Contract $',
        'Customer Payment $', 'Date Pay Recv', 'Plan Gross $', 'Actual Paid PPI $', 'Actual STATZ $',
        'Notes'
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=4, column=col)
        c.font = bold
        c.fill = header_fill
        c.alignment = center
        c.border = border

    seen_contracts = set()
    # 1-indexed columns for currency formatting:
    # Sub PO$(24), Sub Paid$(25), Item Value(26), Contract$(28), Customer Pay$(29),
    # Plan Gross$(31), PPI$(32), STATZ$(33)
    money_cols = {24, 25, 26, 28, 29, 31, 32, 33}

    for clin in clins:
        ack = clin.clinacknowledgment_set.first()
        first_for_contract = clin.contract_id not in seen_contracts
        if first_for_contract:
            seen_contracts.add(clin.contract_id)

        ppi_split_paid = Decimal('0')
        statz_split_paid = Decimal('0')
        if clin.contract_id and first_for_contract:
            ppi_split_paid = (
                ClinSplit.objects.filter(
                    clin__contract_id=clin.contract_id,
                    company_name__iexact='PPI',
                ).aggregate(t=Sum('split_paid'))['t'] or Decimal('0')
            )
            statz_split_paid = (
                ClinSplit.objects.filter(
                    clin__contract_id=clin.contract_id,
                    company_name__iexact='STATZ',
                ).aggregate(t=Sum('split_paid'))['t'] or Decimal('0')
            )

        if clin.contract and clin.contract.status and getattr(clin.contract.status, 'description', '') == 'Canceled':
            status_char = 'X'
        elif clin.contract and clin.contract.date_closed:
            status_char = 'C'
        else:
            status_char = 'O'

        parts = []
        if clin.contract and getattr(clin.contract, 'status', None) and getattr(clin.contract.status, 'description', ''):
            parts.append(clin.contract.status.description)
        if not (ack and ack.po_to_supplier_bool):
            parts.append('PO NOT SENT YET;')
        if not (ack and ack.clin_reply_bool):
            parts.append('SUB REPLY NEEDED;')
        if not (ack and ack.po_to_qar_bool):
            parts.append('PO TO QAR NEEDED;')
        status_text = ' '.join(parts).strip()

        terms = ''
        if clin.contract and getattr(clin.contract, 'special_payment_terms', None):
            spt = clin.contract.special_payment_terms
            terms = getattr(spt, 'terms', None) or getattr(spt, 'code', '') or ''

        qty_uom = f"{float(clin.order_qty):g} {clin.uom or 'ea'}" if clin.order_qty not in (None, '') else ''

        row = [
            status_char,
            (clin.po_number or clin.clin_po_num or (clin.contract.po_number if clin.contract else '')),
            (clin.contract.idiq_contract.contract_number if clin.contract and clin.contract.idiq_contract else ''),
            clin.contract.contract_number if clin.contract else '',
            clin.contract.buyer.description if clin.contract and clin.contract.buyer else '',
            clin.contract.contract_type.description if clin.contract and clin.contract.contract_type else '',
            clin.item_number or '',
            clin.supplier.name if clin.supplier else '',
            clin.supplier.cage_code if clin.supplier else '',
            clin.contract.award_date.strftime('%m/%d/%Y') if clin.contract and clin.contract.award_date else '',
            status_text,
            clin.nsn.nsn_code if clin.nsn else '',
            clin.nsn.description if clin.nsn else '',
            clin.ia or '',
            1 if (ack and ack.po_to_supplier_bool) else 0,
            1 if (ack and ack.clin_reply_bool) else 0,
            1 if (ack and ack.po_to_qar_bool) else 0,
            clin.fob or '',
            clin.supplier_due_date.strftime('%m/%d/%Y') if clin.supplier_due_date else '',
            clin.due_date.strftime('%m/%d/%Y') if clin.due_date else '',
            qty_uom,
            clin.ship_date.strftime('%m/%d/%Y') if clin.ship_date else '',
            float(clin.ship_qty) if clin.ship_qty not in (None, '') else '',
            float(clin.quote_value) if clin.quote_value else '',
            float(clin.paid_amount) if clin.paid_amount else '',
            float(clin.item_value) if clin.item_value else '',
            terms,
            float(clin.contract.contract_value) if (first_for_contract and clin.contract and clin.contract.contract_value) else 0.0,
            float(clin.wawf_payment) if clin.wawf_payment else '',
            clin.wawf_recieved.strftime('%m/%d/%Y') if clin.wawf_recieved else '',
            float(clin.contract.plan_gross) if (first_for_contract and clin.contract and clin.contract.plan_gross is not None) else 0.0,
            float(ppi_split_paid) if (first_for_contract and ppi_split_paid) else 0.0,
            float(statz_split_paid) if (first_for_contract and statz_split_paid) else 0.0,
            int(clin.notes.count())
        ]
        ws.append(row)
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = border

    from openpyxl.utils import get_column_letter
    widths = [6, 10, 18, 18, 14, 12, 8, 24, 10, 12, 26, 12, 24, 8, 10, 10, 10, 8, 10, 10, 10, 10, 10, 12, 12, 12, 12, 14, 14, 14, 14, 14, 14, 30]
    for i in range(1, len(headers) + 1):
        try:
            w = widths[i - 1] if i - 1 < len(widths) else 12
            ws.column_dimensions[get_column_letter(i)].width = w
        except Exception:
            pass

    currency_fmt = '[$$-409]#,##0.00'
    for row in ws.iter_rows(min_row=5, min_col=1, max_col=len(headers), max_row=ws.max_row):
        for idx, cell in enumerate(row, start=1):
            if idx in money_cols and isinstance(cell.value, (int, float)):
                cell.number_format = currency_fmt

    ws.freeze_panes = 'A5'

    from io import BytesIO
    buff = BytesIO()
    wb.save(buff)
    buff.seek(0)

    response = HttpResponse(
        buff.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="contract_log.xlsx"'

    end_time = time.time()
    ExportTiming.objects.create(
        row_count=clins.count(),
        export_time=(end_time - start_time),
        filters_applied=filters_applied
    )

    return response


@conditional_login_required
def get_export_estimate(request):
    """Get estimated export time based on row count."""
    row_count = int(request.GET.get('rows', 0))
    estimated_time = ExportTiming.get_estimated_time(row_count)
    return JsonResponse({
        'estimated_seconds': estimated_time,
        'recent_timings': list(ExportTiming.objects.values('row_count', 'export_time', 'timestamp')[:5])
    })


@conditional_login_required
def open_export_folder(request):
    """Open the export folder in the file explorer."""
    export_folder = os.path.join(settings.MEDIA_ROOT, 'exports')
    os.makedirs(export_folder, exist_ok=True)

    if sys.platform == 'win32':
        os.startfile(export_folder)
    elif sys.platform == 'darwin':
        subprocess.run(['open', export_folder])
    else:
        subprocess.run(['xdg-open', export_folder])

    return JsonResponse({'success': True})
