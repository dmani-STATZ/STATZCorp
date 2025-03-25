from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q
from django.template.loader import render_to_string
from ..models import FolderTracking, Contract
from ..forms import FolderTrackingForm, ContractSearchForm
import json
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import webcolors

def color_to_argb(color):
    """Convert color names or RGB values to ARGB hex format required by openpyxl"""
    try:
        # If it's a hex value
        if color.startswith('#'):
            color = color[1:]
            # Add alpha channel if not present
            if len(color) == 6:
                return 'FF' + color.upper()
            return color.upper()
        
        # Try to convert color name to hex
        rgb = webcolors.name_to_rgb(color.lower())
        hex_color = '%02x%02x%02x' % rgb
        return 'FF' + hex_color.upper()
    except (ValueError, AttributeError):
        # Return white if color is not recognized
        return 'FFFFFFFF'

@login_required
def folder_tracking(request):
    # Get all non-closed records and sort by numeric stack value
    folders = FolderTracking.objects.filter(closed=False).extra(
        select={'stack_num': "CAST(SUBSTRING(stack, 1, CHARINDEX(' -', stack) - 1) AS INTEGER)"}
    ).order_by('stack_num', 'contract__contract_number')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        folders = folders.filter(
            Q(contract__contract_number__icontains=search_query) |
            Q(contract__po_number__icontains=search_query) |
            Q(partial__icontains=search_query) |
            Q(tracking_number__icontains=search_query)
        )

    # Pagination
    paginator = Paginator(folders, 25)  # Show 25 items per page
    page = request.GET.get('page')
    folders = paginator.get_page(page)

    # Get stack colors from model - simplified key format
    stack_colors = {}
    for key, value in FolderTracking.STACK_COLORS.items():
        # Extract the name part (e.g., 'NONE', 'COS', 'PAID', etc.)
        name = key.split(' - ')[1]
        stack_colors[name.lower()] = {
            'bg': value,
            'text': 'white' if value.lower() in ['blue', 'green', 'grey', 'teal', 'red'] else 'black'
        }

    context = {
        'folders': folders,
        'search_form': ContractSearchForm(),
        'search_query': search_query,
        'stack_colors': stack_colors,
    }
    return render(request, 'contracts/folder_tracking.html', context)

@login_required
def search_contracts(request):
    search_query = request.GET.get('q', '')
    search_performed = bool(search_query)
    contracts = []
    
    if search_query:
        contracts = Contract.objects.filter(
            Q(contract_number__icontains=search_query) |
            Q(po_number__icontains=search_query),
            open=True
        ).order_by('contract_number')[:10]  # Limit to 10 results for performance
    
    context = {
        'contracts': contracts,
        'search_performed': search_performed,
    }
    
    # Return direct HTML instead of JSON
    return render(request, 'contracts/includes/contract_search_results.html', context)

@login_required
def add_folder_tracking(request):
    if request.method == 'POST':
        contract_id = request.POST.get('contract')
        contract = get_object_or_404(Contract, id=contract_id)
        
        # Check if a folder already exists for this contract
        if FolderTracking.objects.filter(contract=contract, closed=False).exists():
            return JsonResponse({
                'status': 'error',
                'message': 'A folder already exists for this contract'
            })
        
        # Create new folder with default values
        folder = FolderTracking.objects.create(
            contract=contract,
            stack='0 - NONE',  # Default to first stack
            added_by=request.user,
            created_by=request.user,
            modified_by=request.user
        )
        return JsonResponse({'status': 'success'})
        
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

@login_required
def close_folder_tracking(request, pk):
    folder = get_object_or_404(FolderTracking, pk=pk)
    folder.close_record(request.user)
    return JsonResponse({'status': 'success'})

@login_required
def toggle_highlight(request, pk):
    folder = get_object_or_404(FolderTracking, pk=pk)
    folder.toggle_highlight(request.user)
    return JsonResponse({'status': 'success', 'highlighted': folder.highlight})

@login_required
def export_folder_tracking(request):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="folder_tracking_{datetime.now().strftime("%Y%m%d")}.xlsx"'

    # Create workbook and select active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Folder Tracking"

    # Define headers
    headers = ['Stack', 'Contract', 'PO', 'Partial', 'RTS Email', 'QB INV', 'WAWF', 'WAWF QAR',
               'VSM SCN', 'SIR SCN', 'Tracking', 'Tracking #', 'Sort Data', 'Note']

    # Write headers with styling
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Initialize dictionary to track maximum width of each column
    max_lengths = {i: len(str(header)) for i, header in enumerate(headers)}
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Get folders and sort them
    folders = FolderTracking.objects.filter(closed=False).extra(
        select={'stack_num': "CAST(SUBSTRING(stack, 1, CHARINDEX(' -', stack) - 1) AS INTEGER)"}
    ).order_by('stack_num', 'contract__contract_number')

    # Write data with styling
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Define boolean fields (these will be center-aligned)
    boolean_fields = {'RTS Email', 'WAWF', 'WAWF QAR'}

    for row_idx, folder in enumerate(folders, 2):
        # Get stack color
        stack_color = folder.stack_color
        if not stack_color:
            stack_color = 'FFFFFF'  # Default to white if no color specified
        
        # Convert color to ARGB format
        argb_color = color_to_argb(stack_color)
        
        # Create fill style for the stack column
        stack_fill = PatternFill(start_color=argb_color, end_color=argb_color, fill_type="solid")

        # Write row data with styling
        row_data = [
            folder.stack,
            folder.contract.contract_number,
            folder.contract.po_number,
            folder.partial or '',
            'Yes' if folder.rts_email else 'No',
            folder.qb_inv or '',
            'Yes' if folder.wawf else 'No',
            'Yes' if folder.wawf_qar else 'No',
            folder.vsm_scn or '',
            folder.sir_scn or '',
            folder.tracking or '',
            folder.tracking_number or '',
            folder.sort_data or '',
            folder.note or ''
        ]

        # Update maximum lengths
        for i, value in enumerate(row_data):
            max_lengths[i] = max(max_lengths[i], len(str(value or '')))

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Apply basic styling to all cells
            cell.border = thin_border
            cell.font = Font(color="000000")  # Black text for all cells
            
            # Set alignment based on field type
            if headers[col_idx - 1] in boolean_fields:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Apply color only to stack column
            if col_idx == 1:  # Stack column
                cell.fill = stack_fill
            else:
                cell.fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")  # White background
            
            # Apply highlight to non-stack columns if needed
            if folder.highlight and col_idx > 1:
                cell.fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")  # ARGB yellow

    # Adjust column widths based on content
    for i, max_length in max_lengths.items():
        column = get_column_letter(i + 1)
        # Set width with some padding and maximum width limit
        adjusted_width = min(max_length + 2, 50)  # Add 2 for padding, cap at 50
        ws.column_dimensions[column].width = adjusted_width

    # Save workbook
    wb.save(response)
    return response

@login_required
def update_folder_field(request, pk):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})
    
    folder = get_object_or_404(FolderTracking, pk=pk)
    field = request.POST.get('field')
    value = request.POST.get('value')
    
    if field not in ['stack', 'partial', 'rts_email', 'qb_inv', 'wawf', 'wawf_qar', 
                    'vsm_scn', 'sir_scn', 'tracking', 'tracking_number', 'note']:
        return JsonResponse({'status': 'error', 'message': 'Invalid field'})
    
    # Handle boolean fields
    if field in ['rts_email', 'wawf', 'wawf_qar']:
        value = value.lower() == 'true'
    
    # Update the field value
    setattr(folder, field, value)
    
    # Update audit fields
    folder.modified_by = request.user
    # modified_on will be automatically updated by Django's auto_now=True
    
    folder.save()
    
    return JsonResponse({
        'status': 'success',
        'value': value if field not in ['rts_email', 'wawf', 'wawf_qar'] else bool(value)
    }) 