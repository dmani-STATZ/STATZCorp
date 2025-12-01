from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render
from django.utils import timezone
import csv
import re
from io import TextIOWrapper
from pathlib import Path

from openpyxl import load_workbook

from ..models import Supplier


def _is_superuser(user):
    return user.is_superuser


def _extract_name_and_url(raw_name):
    """
    Handles several formats:
    - "Display Text" (no URL) -> (display, None)
    - "Display Text|https://..." -> split on first pipe
    - "Display Text,https://..." from CSV hyperlink export
    - Anything with http substring -> grab http.. as url, left part as name
    """
    if not raw_name:
        return "", None
    text = raw_name.strip().strip('"')
    # pipe-delimited (common Excel hyperlink export)
    if "|" in text:
        left, right = text.split("|", 1)
        return left.strip(), right.strip()
    # comma-delimited hyperlink
    if "," in text and "http" in text:
        parts = text.split(",", 1)
        left = parts[0].strip()
        url_part = parts[1].strip()
        return left, url_part if url_part.lower().startswith("http") else None
    # generic http search
    match = re.search(r"(https?://\\S+)", text)
    if match:
        url = match.group(1)
        name_part = text[: match.start()].strip(" ,")
        return name_part or text, url
    return text, None


@login_required
@user_passes_test(_is_superuser)
def supplier_admin_tools(request):
    """
    Admin page to bulk update supplier.files_url from CSV.
    Supports:
      - Simple CSV with columns: supplier_id (or name), files_url
      - SharePoint export CSV with columns: Name, Item Type, Modified, etc.
        * Only rows with Item Type == Folder
        * Name may contain hyperlink; we extract URL from the Name field.
      - SharePoint export XLSX: Name cell hyperlink is read directly
    """
    summary = {}
    rows = []

    if request.method == "POST":
        upload = request.FILES.get("files_csv") or request.FILES.get("sharepoint_csv")
        if not upload:
            messages.error(request, "Please choose a CSV file to upload.")
        else:
            try:
                updated, missing, errors, skipped = 0, 0, 0, 0
                ext = Path(upload.name or "").suffix.lower()

                def process_row(idx, supplier_id, name, files_url, item_type):
                    nonlocal updated, missing, errors, skipped
                    is_sharepoint = bool(item_type)
                    if is_sharepoint and item_type.lower() != "folder":
                        skipped += 1
                        return

                    # If SharePoint export, derive url from Name, ignore files_url column if empty
                    if is_sharepoint:
                        name_text, url = _extract_name_and_url(name)
                        name = name_text
                        if url:
                            files_url = url
                    if not files_url:
                        skipped += 1
                        return

                    supplier = None
                    if supplier_id:
                        try:
                            supplier = Supplier.objects.filter(pk=int(supplier_id)).first()
                        except ValueError:
                            supplier = None
                    if not supplier and name:
                        supplier = Supplier.objects.filter(name__iexact=name).first()

                    if not supplier:
                        missing += 1
                        rows.append({"row": idx, "status": "missing", "supplier": supplier_id or name or "-", "url": files_url})
                        return

                    try:
                        supplier.files_url = files_url or None
                        supplier.modified_by = request.user
                        supplier.modified_on = timezone.now()
                        supplier.save(update_fields=["files_url", "modified_by", "modified_on"])
                        updated += 1
                        rows.append({"row": idx, "status": "updated", "supplier": supplier.name, "url": files_url or "(cleared)"})
                    except Exception:
                        errors += 1
                        rows.append({"row": idx, "status": "error", "supplier": supplier.name, "url": files_url or "(cleared)"})

                if ext == ".xlsx":
                    wb = load_workbook(upload, read_only=True, data_only=True)
                    ws = wb.active
                    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
                    header_map = {h.lower(): idx for idx, h in enumerate(headers)}
                    def get_cell(row, key):
                        idx = header_map.get(key.lower())
                        if idx is None or idx >= len(row):
                            return ""
                        cell = row[idx]
                        return cell
                    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                        name_cell = get_cell(row, "name")
                        name_val = name_cell.value or ""
                        url = name_cell.hyperlink.target if name_cell and name_cell.hyperlink else ""
                        item_type_cell = get_cell(row, "Item Type")
                        item_type_val = item_type_cell.value if hasattr(item_type_cell, "value") else ""
                        supplier_id_val = ""
                        files_url_val = url or ""
                        process_row(idx, supplier_id_val, name_val, files_url_val, item_type_val or "")
                else:
                    decoded = TextIOWrapper(upload, encoding="utf-8-sig")
                    reader = csv.DictReader(decoded)
                    for idx, row in enumerate(reader, start=1):
                        item_type = (row.get("Item Type") or row.get("item type") or row.get("item_type") or "").strip()
                        supplier_id = (row.get("supplier_id") or row.get("id") or "").strip()
                        name = (row.get("name") or row.get("Name") or "").strip()
                        files_url = (row.get("files_url") or row.get("url") or "").strip()
                        process_row(idx, supplier_id, name, files_url, item_type)

                total = updated + missing + errors + skipped
                summary = {"updated": updated, "missing": missing, "errors": errors, "skipped": skipped, "total": total}
                messages.success(request, f"Processed {total} rows. Updated {updated}, Missing {missing}, Errors {errors}, Skipped {skipped}.")
            except Exception as exc:
                messages.error(request, f"Upload failed: {exc}")

    context = {
        "summary": summary,
        "rows": rows,
    }
    return render(request, "contracts/admin_tools.html", context)
