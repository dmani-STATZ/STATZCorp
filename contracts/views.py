from django.shortcuts import render, get_object_or_404, redirect
from STATZWeb.decorators import conditional_login_required
from django.views.generic import TemplateView, DetailView, UpdateView, CreateView, ListView, DeleteView
from django.db.models import Count, Sum, Q, F
from django.utils import timezone
from django.utils.safestring import mark_safe
from django.http import JsonResponse, HttpResponseRedirect
from datetime import timedelta, datetime
from pdfminer.high_level import extract_pages, extract_text
import calendar
from django.contrib import messages
from django.contrib.contenttypes.models import ContentType
from django.urls import reverse_lazy, reverse
from django.views.decorators.http import require_http_methods
from django.utils.decorators import method_decorator
import json
import re
import tempfile
import os
import PyPDF2
import pytesseract
from PIL import Image
import pdf2image
import io
import logging
import fitz  # PyMuPDF
from django.conf import settings
import sys
import subprocess
import csv
# Try to import openpyxl, but make it optional
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
from decimal import Decimal
from django.db.models import Q, Sum, Count, F, Value, CharField
from django.db.models.functions import Concat
from django.contrib.auth.decorators import login_required
from django.template.loader import render_to_string
from django.core.exceptions import PermissionDenied

from .models import (
    Contract, Clin, ClinFinance, Supplier, Nsn, ClinAcknowledgment, 
    Note, Reminder, Address, Contact, AcknowledgementLetter
)
from .forms import (
    NsnForm, SupplierForm, ContractForm, ContractCloseForm, ContractCancelForm,
    ClinForm, ClinFinanceForm, NoteForm, ReminderForm, ClinAcknowledgmentForm,
    AddressForm, ContactForm
)

# Configure logger
logger = logging.getLogger(__name__)

# Create your views here.
@method_decorator(conditional_login_required, name='dispatch')
class ContractDetailView(DetailView):
    model = Contract
    template_name = 'contracts/contract_detail.html'
    context_object_name = 'contract'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        contract = self.get_object()
        clins = contract.clin_set.all().select_related(
            'clin_type', 'supplier', 'nsn'
        )
        context['clins'] = clins
        
        # Use the notes attribute directly instead of note_set
        context['contract_notes'] = contract.notes.all().order_by('-created_on')
        
        # Get the default selected CLIN (type=1) or first CLIN if no type 1 exists
        context['selected_clin'] = clins.filter(clin_type_id=1).first() or clins.first()
        if context['selected_clin']:
            # Use the notes attribute directly instead of note_set
            context['clin_notes'] = context['selected_clin'].notes.all().order_by('-created_on')
            try:
                context['acknowledgment'] = context['selected_clin'].clinacknowledgment_set.first()
            except:
                context['acknowledgment'] = None
        else:
            context['clin_notes'] = []
            context['acknowledgment'] = None
            
        return context


@method_decorator(conditional_login_required, name='dispatch')
class ClinDetailView(DetailView):
    model = Clin
    template_name = 'contracts/clin_detail.html'
    context_object_name = 'clin'

    def get_queryset(self):
        return super().get_queryset().select_related(
            'contract',
            'clin_type',
            'supplier',
            'nsn',
            'clin_finance',
            'clin_finance__special_payment_terms'
        )


@conditional_login_required
def contract_search(request):
    query = request.GET.get('q', '')
    if len(query) < 3:
        return JsonResponse([], safe=False)

    # Search by contract number, last 6 characters, or contract's PO number
    contracts = Contract.objects.filter(
        Q(contract_number__icontains=query) |
        (Q(contract_number__iendswith=query[-6:]) if len(query) >= 6 else Q()) |
        Q(po_number__icontains=query)  # Search Contract's po_number field
    ).values(
        'id', 
        'contract_number',
        'po_number'  # Include contract's PO number in results
    ).order_by('contract_number')[:10]
    
    # Format the results
    results = []
    for contract in contracts:
        contract_data = {
            'id': contract['id'],
            'contract_number': contract['contract_number'],
            'po_numbers': []
        }
        
        # Add contract's PO number if it exists and matches the query
        if contract['po_number'] and query.lower() in contract['po_number'].lower():
            contract_data['po_numbers'].append(contract['po_number'])
        
        results.append(contract_data)
    
    return JsonResponse(results, safe=False)


@method_decorator(conditional_login_required, name='dispatch')
class NsnUpdateView(UpdateView):
    model = Nsn
    template_name = 'contracts/nsn_edit.html'
    context_object_name = 'nsn'
    form_class = NsnForm
    
    def get_success_url(self):
        next_url = self.request.GET.get('next')
        if next_url:
            return next_url
        return reverse_lazy('contracts:contracts_dashboard')


@method_decorator(conditional_login_required, name='dispatch')
class SupplierUpdateView(UpdateView):
    model = Supplier
    template_name = 'contracts/supplier_edit.html'
    context_object_name = 'supplier'
    form_class = SupplierForm
    
    def get_success_url(self):
        next_url = self.request.GET.get('next')
        if next_url:
            return next_url
        return reverse_lazy('contracts:contracts_dashboard')


@conditional_login_required
def get_clin_notes(request, clin_id):
    clin = get_object_or_404(Clin, id=clin_id)
    notes = clin.clinnote_set.all().order_by('-created_on')
    notes_data = [{
        'note': note.note,
        'created_by': str(note.created_by),
        'created_on': note.created_on.strftime("%b %d, %Y %H:%M")
    } for note in notes]
    return JsonResponse({'notes': notes_data})


@conditional_login_required
@require_http_methods(["POST"])
def toggle_clin_acknowledgment(request, clin_id):
    try:
        clin = Clin.objects.get(id=clin_id)
        data = json.loads(request.body)
        field = data.get('field')
        
        # Get or create acknowledgment
        acknowledgment, created = ClinAcknowledgment.objects.get_or_create(clin=clin)
        
        # Toggle the field
        current_value = getattr(acknowledgment, field)
        new_value = not current_value
        
        # Update the boolean field
        setattr(acknowledgment, field, new_value)
        
        # Update the corresponding date and user fields
        field_base = field.replace('_bool', '')
        date_field = f"{field_base}_date"
        user_field = f"{field_base}_user"
        
        if new_value:
            setattr(acknowledgment, date_field, timezone.now())
            setattr(acknowledgment, user_field, request.user.username)
        else:
            setattr(acknowledgment, date_field, None)
            setattr(acknowledgment, user_field, None)
        
        acknowledgment.save()
        
        return JsonResponse({
            'status': new_value,
            'date': getattr(acknowledgment, date_field).isoformat() if new_value else None
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

# Contract Management Views
@method_decorator(conditional_login_required, name='dispatch')
class ContractCreateView(CreateView):
    model = Contract
    form_class = ContractForm
    template_name = 'contracts/contract_form.html'
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.modified_by = self.request.user
        form.instance.open = True
        response = super().form_valid(form)
        messages.success(self.request, f'Contract {form.instance.contract_number} created successfully.')
        return response
    
    def get_success_url(self):
        return reverse('contracts:contract_detail', kwargs={'pk': self.object.pk})

@method_decorator(conditional_login_required, name='dispatch')
class ContractUpdateView(UpdateView):
    model = Contract
    form_class = ContractForm
    template_name = 'contracts/contract_form.html'
    
    def form_valid(self, form):
        form.instance.modified_by = self.request.user
        response = super().form_valid(form)
        messages.success(self.request, f'Contract {form.instance.contract_number} updated successfully.')
        return response
    
    def get_success_url(self):
        return reverse('contracts:contract_detail', kwargs={'pk': self.object.pk})

@method_decorator(conditional_login_required, name='dispatch')
class ContractCloseView(UpdateView):
    model = Contract
    form_class = ContractCloseForm
    template_name = 'contracts/contract_close_form.html'
    
    def form_valid(self, form):
        form.instance.modified_by = self.request.user
        if not form.instance.open:
            form.instance.date_closed = timezone.now()
        response = super().form_valid(form)
        status = "closed" if not form.instance.open else "reopened"
        messages.success(self.request, f'Contract {form.instance.contract_number} {status} successfully.')
        return response
    
    def get_success_url(self):
        return reverse('contracts:contract_detail', kwargs={'pk': self.object.pk})

@method_decorator(conditional_login_required, name='dispatch')
class ContractCancelView(UpdateView):
    model = Contract
    form_class = ContractCancelForm
    template_name = 'contracts/contract_cancel_form.html'
    
    def form_valid(self, form):
        form.instance.modified_by = self.request.user
        if form.instance.cancelled:
            form.instance.date_canceled = timezone.now()
        response = super().form_valid(form)
        status = "cancelled" if form.instance.cancelled else "uncancelled"
        messages.success(self.request, f'Contract {form.instance.contract_number} {status} successfully.')
        return response
    
    def get_success_url(self):
        return reverse('contracts:contract_detail', kwargs={'pk': self.object.pk})

# CLIN Management Views
@method_decorator(conditional_login_required, name='dispatch')
class ClinCreateView(CreateView):
    model = Clin
    form_class = ClinForm
    template_name = 'contracts/clin_form.html'
    
    def get_initial(self):
        initial = super().get_initial()
        if 'contract_id' in self.kwargs:
            initial['contract'] = self.kwargs['contract_id']
        return initial
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.modified_by = self.request.user
        
        # Create ClinFinance instance if it doesn't exist
        response = super().form_valid(form)
        if not hasattr(form.instance, 'clin_finance') or not form.instance.clin_finance:
            clin_finance = ClinFinance.objects.create(
                created_by=self.request.user,
                modified_by=self.request.user
            )
            form.instance.clin_finance = clin_finance
            form.instance.save()
        
        # Create ClinAcknowledgment instance if it doesn't exist
        if not ClinAcknowledgment.objects.filter(clin=form.instance).exists():
            ClinAcknowledgment.objects.create(
                clin=form.instance,
                created_by=self.request.user,
                modified_by=self.request.user
            )
        
        messages.success(self.request, f'CLIN created successfully.')
        return response
    
    def get_success_url(self):
        return reverse('contracts:clin_detail', kwargs={'pk': self.object.pk})

@method_decorator(conditional_login_required, name='dispatch')
class ClinUpdateView(UpdateView):
    model = Clin
    form_class = ClinForm
    template_name = 'contracts/clin_form.html'
    
    def form_valid(self, form):
        form.instance.modified_by = self.request.user
        response = super().form_valid(form)
        messages.success(self.request, f'CLIN updated successfully.')
        return response
    
    def get_success_url(self):
        return reverse('contracts:clin_detail', kwargs={'pk': self.object.pk})

@method_decorator(conditional_login_required, name='dispatch')
class ClinFinanceUpdateView(UpdateView):
    model = ClinFinance
    form_class = ClinFinanceForm
    template_name = 'contracts/clin_finance_form.html'
    
    def form_valid(self, form):
        form.instance.modified_by = self.request.user
        response = super().form_valid(form)
        messages.success(self.request, f'CLIN Finance information updated successfully.')
        return response
    
    def get_success_url(self):
        clin = Clin.objects.get(clin_finance=self.object)
        return reverse('contracts:clin_detail', kwargs={'pk': clin.pk})

@method_decorator(conditional_login_required, name='dispatch')
class ClinAcknowledgmentUpdateView(UpdateView):
    model = ClinAcknowledgment
    form_class = ClinAcknowledgmentForm
    template_name = 'contracts/clin_acknowledgment_form.html'
    
    def form_valid(self, form):
        form.instance.modified_by = self.request.user
        
        # Update user fields based on checkboxes
        if form.instance.po_to_supplier_bool and not form.instance.po_to_supplier_user:
            form.instance.po_to_supplier_user = self.request.user
            form.instance.po_to_supplier_date = timezone.now()
            
        if form.instance.clin_reply_bool and not form.instance.clin_reply_user:
            form.instance.clin_reply_user = self.request.user
            form.instance.clin_reply_date = timezone.now()
            
        if form.instance.po_to_qar_bool and not form.instance.po_to_qar_user:
            form.instance.po_to_qar_user = self.request.user
            form.instance.po_to_qar_date = timezone.now()
            
        response = super().form_valid(form)
        messages.success(self.request, f'CLIN Acknowledgment updated successfully.')
        return response
    
    def get_success_url(self):
        return reverse('contracts:clin_detail', kwargs={'pk': self.object.clin.pk})

# Note Management Views
@conditional_login_required
def add_note(request, content_type_id, object_id):
    content_type = get_object_or_404(ContentType, id=content_type_id)
    content_object = get_object_or_404(content_type.model_class(), id=object_id)
    
    if request.method == 'POST':
        form = NoteForm(request.POST)
        if form.is_valid():
            note = form.save(commit=False)
            note.content_type = content_type
            note.object_id = object_id
            note.created_by = request.user
            note.modified_by = request.user
            note.save()
            
            # Determine the redirect URL based on the content type
            if content_type.model == 'contract':
                return redirect('contracts:contract_detail', pk=object_id)
            elif content_type.model == 'clin':
                return redirect('contracts:clin_detail', pk=object_id)
            else:
                return redirect('contracts:contracts_dashboard')
    else:
        form = NoteForm()
    
    context = {
        'form': form,
        'content_type_id': content_type_id,
        'object_id': object_id,
        'content_object': content_object,
    }
    return render(request, 'contracts/note_form.html', context)

@conditional_login_required
def delete_note(request, note_id):
    note = get_object_or_404(Note, id=note_id)
    content_object = note.content_object
    content_type = note.content_type
    
    # Check if the user is the creator of the note
    if note.created_by != request.user and not request.user.is_staff:
        messages.error(request, "You don't have permission to delete this note.")
        if content_type.model == 'contract':
            return redirect('contracts:contract_detail', pk=content_object.id)
        elif content_type.model == 'clin':
            return redirect('contracts:clin_detail', pk=content_object.id)
        else:
            return redirect('contracts:contracts_dashboard')
    
    if request.method == 'POST':
        note.delete()
        messages.success(request, "Note deleted successfully.")
        
        # Determine the redirect URL based on the content type
        if content_type.model == 'contract':
            return redirect('contracts:contract_detail', pk=content_object.id)
        elif content_type.model == 'clin':
            return redirect('contracts:clin_detail', pk=content_object.id)
        else:
            return redirect('contracts:contracts_dashboard')
    
    context = {
        'note': note,
        'content_object': content_object,
    }
    return render(request, 'contracts/note_confirm_delete.html', context)

# Reminder Management Views
@conditional_login_required
def add_reminder(request, note_id=None):
    note = None
    if note_id:
        note = get_object_or_404(Note, id=note_id)
    
    if request.method == 'POST':
        form = ReminderForm(request.POST)
        if form.is_valid():
            reminder = form.save(commit=False)
            reminder.reminder_user = request.user
            reminder.note = note
            reminder.save()
            
            messages.success(request, "Reminder created successfully.")
            
            # Redirect based on whether the reminder is associated with a note
            if note:
                content_object = note.content_object
                content_type = note.content_type
                if content_type.model == 'contract':
                    return redirect('contracts:contract_detail', pk=content_object.id)
                elif content_type.model == 'clin':
                    return redirect('contracts:clin_detail', pk=content_object.id)
            
            return redirect('contracts:reminders_list')
    else:
        initial_data = {}
        if note:
            initial_data['reminder_text'] = f"Follow up on: {note.note[:100]}..."
        form = ReminderForm(initial=initial_data)
    
    context = {
        'form': form,
        'note': note,
    }
    return render(request, 'contracts/reminder_form.html', context)

@method_decorator(conditional_login_required, name='dispatch')
class ReminderListView(ListView):
    model = Reminder
    template_name = 'contracts/reminders_list.html'
    context_object_name = 'reminders'
    
    def get_queryset(self):
        # Get reminders for the current user
        queryset = Reminder.objects.filter(
            reminder_user=self.request.user
        ).order_by('reminder_completed', 'reminder_date')
        
        # Filter by completion status if specified
        status = self.request.GET.get('status')
        if status == 'completed':
            queryset = queryset.filter(reminder_completed=True)
        elif status == 'pending':
            queryset = queryset.filter(Q(reminder_completed=False) | Q(reminder_completed=None))
        
        # Filter by date range if specified
        date_filter = self.request.GET.get('date_filter')
        today = timezone.now().date()
        if date_filter == 'today':
            queryset = queryset.filter(reminder_date__date=today)
        elif date_filter == 'week':
            week_start = today - timedelta(days=today.weekday())
            week_end = week_start + timedelta(days=6)
            queryset = queryset.filter(reminder_date__date__range=[week_start, week_end])
        elif date_filter == 'month':
            month_start = today.replace(day=1)
            if today.month == 12:
                month_end = today.replace(year=today.year+1, month=1, day=1) - timedelta(days=1)
            else:
                month_end = today.replace(month=today.month+1, day=1) - timedelta(days=1)
            queryset = queryset.filter(reminder_date__date__range=[month_start, month_end])
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status'] = self.request.GET.get('status', 'all')
        context['date_filter'] = self.request.GET.get('date_filter', 'all')
        
        # Add counts for different statuses
        context['total_count'] = Reminder.objects.filter(reminder_user=self.request.user).count()
        context['completed_count'] = Reminder.objects.filter(
            reminder_user=self.request.user, 
            reminder_completed=True
        ).count()
        context['pending_count'] = Reminder.objects.filter(
            reminder_user=self.request.user
        ).filter(
            Q(reminder_completed=False) | Q(reminder_completed=None)
        ).count()
        
        # Add counts for overdue reminders
        context['overdue_count'] = Reminder.objects.filter(
            reminder_user=self.request.user,
            reminder_date__lt=timezone.now()
        ).filter(
            Q(reminder_completed=False) | Q(reminder_completed=None)
        ).count()
        
        return context

@conditional_login_required
def toggle_reminder_completion(request, reminder_id):
    reminder = get_object_or_404(Reminder, id=reminder_id)
    
    # Check if the user is the owner of the reminder
    if reminder.reminder_user != request.user and not request.user.is_staff:
        messages.error(request, "You don't have permission to update this reminder.")
        return redirect('contracts:reminders_list')
    
    # Toggle the completion status
    reminder.reminder_completed = not reminder.reminder_completed
    if reminder.reminder_completed:
        reminder.reminder_completed_date = timezone.now()
        reminder.reminder_completed_user = request.user
    else:
        reminder.reminder_completed_date = None
        reminder.reminder_completed_user = None
    
    reminder.save()
    
    # Redirect back to the referring page or the reminders list
    next_url = request.GET.get('next')
    if next_url:
        return redirect(next_url)
    return redirect('contracts:reminders_list')

@conditional_login_required
def delete_reminder(request, reminder_id):
    reminder = get_object_or_404(Reminder, id=reminder_id)
    
    # Check if the user is the owner of the reminder
    if reminder.reminder_user != request.user and not request.user.is_staff:
        messages.error(request, "You don't have permission to delete this reminder.")
        return redirect('contracts:reminders_list')
    
    if request.method == 'POST':
        reminder.delete()
        messages.success(request, "Reminder deleted successfully.")
        
        # Redirect back to the referring page or the reminders list
        next_url = request.GET.get('next')
        if next_url:
            return redirect(next_url)
        return redirect('contracts:reminders_list')
    
    context = {
        'reminder': reminder,
    }
    return render(request, 'contracts/reminder_confirm_delete.html', context)

# Contract Lifecycle Dashboard
@method_decorator(conditional_login_required, name='dispatch')
class ContractLifecycleDashboardView(TemplateView):
    template_name = 'contracts/contract_lifecycle_dashboard.html'

    def get_contracts(self):
        # Get the last 20 contracts entered that have cancelled=False
        last_20_contracts = Contract.objects.filter(
                cancelled=False
            ).prefetch_related(
                'clin_set',
                'clin_set__clin_finance',
                'clin_set__supplier'
            ).order_by('-created_on')[:20]

        # Prepare the data for rendering or serialization
        contracts_data = []
        for contract in last_20_contracts:
            # Get the first CLIN with clin_type_id=1 for this contract
            main_clin = contract.clin_set.filter(clin_type_id=1).first()
            if main_clin and main_clin.clin_finance and main_clin.supplier:
                contracts_data.append({
                    'id': contract.id,
                    'tab_num': contract.tab_num,
                    'po_number': contract.po_number,
                    'contract_number': contract.contract_number,
                    'supplier_name': main_clin.supplier.name,
                    'contract_value': main_clin.clin_finance.contract_value,
                    'award_date': contract.award_date,
                    'due_date': contract.due_date,
                    'cancelled': contract.cancelled,
                    'open': contract.open,
                })

        return contracts_data
        
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        now = timezone.now()
        
        # Time periods
        this_week_start = now - timedelta(days=now.weekday())
        this_week_end = this_week_start + timedelta(days=6)
        last_week_start = this_week_start - timedelta(weeks=1)
        last_week_end = last_week_start + timedelta(days=6)

                # Calculate month boundaries
        this_month_start = now.replace(day=1)
        this_month_end = now.replace(day=calendar.monthrange(now.year, now.month)[1])
        
        # Calculate last month
        if now.month == 1:
            last_month_start = now.replace(year=now.year-1, month=12, day=1)
            last_month_end = now.replace(year=now.year-1, month=12, day=31)
        else:
            last_month_start = now.replace(month=now.month-1, day=1)
            last_month_end = now.replace(month=now.month-1, day=calendar.monthrange(now.year, now.month-1)[1])
        
        # Calculate quarter starts and ends
        current_quarter = (now.month - 1) // 3
        this_quarter_start = now.replace(month=current_quarter * 3 + 1, day=1)
        this_quarter_end = now.replace(
            month=min(12, (current_quarter + 1) * 3),
            day=calendar.monthrange(now.year, min(12, (current_quarter + 1) * 3))[1]
        )

        if current_quarter == 0:  # If we're in Q1
            last_quarter_start = now.replace(year=now.year - 1, month=10, day=1)
            last_quarter_end = now.replace(year=now.year - 1, month=12, day=31)
        else:
            last_quarter_start = now.replace(month=((current_quarter - 1) * 3) + 1, day=1)
            last_quarter_month = min(12, (current_quarter) * 3)
            last_quarter_end = now.replace(
                month=last_quarter_month,
                day=calendar.monthrange(now.year, last_quarter_month)[1]
            )

        this_year_start = now.replace(month=1, day=1)
        this_year_end = now.replace(month=12, day=31)
        last_year_start = this_year_start.replace(year=this_year_start.year-1)
        last_year_end = last_year_start.replace(month=12, day=31)

        # Helper function to get stats for a time period
        def get_period_stats(start_date, end_date=None):
            if not end_date:
                end_date = now

            past_contracts = Contract.objects.filter(due_date__range=(start_date, end_date),cancelled=False)
            contracts = Contract.objects.filter(award_date__range=(start_date, end_date),cancelled=False)
            clins = Clin.objects.filter(contract__award_date__range=(start_date, end_date),contract__cancelled=False)
            
            return {
                'contracts_due': past_contracts.distinct().count(),
                'contracts_due_late': past_contracts.filter(due_date_late=True).distinct().count(),
                'contracts_due_ontime': past_contracts.filter(due_date_late=False).distinct().count(),
                'new_contract_value': clins.aggregate(total=Sum('clin_finance__contract_value'))['total'] or 0,
                'new_contracts': contracts.distinct().count(),
                'date_range': mark_safe(f"{start_date.strftime('%Y/%m/%d')} to<br>{end_date.strftime('%Y/%m/%d')}"),
            }
        
                # Get stats for each time period
        periods = {
            'this_week': get_period_stats(this_week_start, this_week_end),
            'last_week': get_period_stats(last_week_start, last_week_end),
            'this_month': get_period_stats(this_month_start,this_month_end),
            'last_month': get_period_stats(last_month_start, last_month_end),
            'this_quarter': get_period_stats(this_quarter_start, this_quarter_end),
            'last_quarter': get_period_stats(last_quarter_start, last_quarter_end),
            'this_year': get_period_stats(this_year_start, this_year_end),
            'last_year': get_period_stats(last_year_start, last_year_end),
        }

        context['periods'] = periods
        context['contracts'] = self.get_contracts()
        # Get all active contracts (not cancelled and not closed)
        active_contracts = Contract.objects.filter(
            Q(cancelled=False) & (Q(open=True) | Q(open=None)))
        
        # Contracts by stage
        context['new_contracts'] = active_contracts.filter(
            award_date__gte=timezone.now() - timedelta(days=30)
        ).count()
        
        # Contracts with CLINs that have acknowledgments pending
        contracts_with_pending_acks = Contract.objects.filter(
            clin__clinacknowledgment__po_to_supplier_bool=True,
            clin__clinacknowledgment__clin_reply_bool=False
        ).distinct().count()
        context['pending_acknowledgment'] = contracts_with_pending_acks
        
        # Contracts with CLINs that are in production (acknowledged but not shipped)
        contracts_in_production = Contract.objects.filter(
            clin__clinacknowledgment__clin_reply_bool=True,
            clin__ship_date=None
        ).distinct().count()
        context['in_production'] = contracts_in_production
        
        # Contracts with CLINs that are shipped but not paid
        contracts_shipped_not_paid = Contract.objects.filter(
            clin__ship_date__isnull=False,
            clin__clin_finance__paid_date=None
        ).distinct().count()
        context['shipped_not_paid'] = contracts_shipped_not_paid
        
        # Contracts with all CLINs paid
        contracts_all_paid = Contract.objects.annotate(
            total_clins=Count('clin'),
            paid_clins=Count('clin', filter=Q(clin__clin_finance__paid_date__isnull=False))
        ).filter(
            total_clins=F('paid_clins'),
            total_clins__gt=0
        ).count()
        context['fully_paid'] = contracts_all_paid
        
        # Contracts with upcoming due dates
        context['due_soon'] = active_contracts.filter(
            due_date__range=[timezone.now(), timezone.now() + timedelta(days=14)]
        ).count()
        
        # Contracts that are past due
        context['past_due'] = active_contracts.filter(
            due_date__lt=timezone.now(),
            due_date_late=True
        ).count()
        
        # User's reminders
        context['pending_reminders'] = Reminder.objects.filter(
            reminder_user=self.request.user,
            reminder_completed=False,
            reminder_date__lte=timezone.now() + timedelta(days=7)
        ).order_by('reminder_date')[:5]

 
        return context

# Supplier Communication
@conditional_login_required
def generate_acknowledgement_letter(request, clin_id):
    clin = get_object_or_404(Clin, id=clin_id)
    
    # Check if an acknowledgement letter already exists
    existing_letter = AcknowledgementLetter.objects.filter(clin=clin).first()
    
    if request.method == 'POST':
        if existing_letter:
            # Update existing letter
            existing_letter.modified_by = request.user
            existing_letter.letter_date = timezone.now()
            existing_letter.save()
        else:
            # Create new letter with default values
            supplier = clin.supplier
            contact = supplier.contact if supplier else None
            address = contact.address if contact else None
            
            letter = AcknowledgementLetter(
                clin=clin,
                created_by=request.user,
                modified_by=request.user,
                letter_date=timezone.now(),
                salutation=contact.salutation if contact else '',
                addr_fname=contact.name.split()[0] if contact and contact.name else '',
                addr_lname=' '.join(contact.name.split()[1:]) if contact and contact.name and len(contact.name.split()) > 1 else '',
                supplier=supplier.name if supplier else '',
                st_address=address.address_line_1 if address else '',
                city=address.city if address else '',
                state=address.state if address else '',
                zip=address.zip if address else '',
                po=clin.po_number or '',
                po_ext=clin.po_num_ext or '',
                contract_num=clin.contract.contract_number if clin.contract else '',
                fat_plt_due_date=clin.due_date,
                supplier_due_date=clin.supplier_due_date,
                dpas_priority='DO-A3',  # Default value, could be made configurable
                statz_contact=request.user.get_full_name(),
                statz_contact_title='Contract Manager',  # Default value, could be made configurable
                statz_contact_phone='555-123-4567',  # Default value, could be made configurable
                statz_contact_email=request.user.email,
            )
            letter.save()
        
        messages.success(request, "Acknowledgement letter generated successfully.")
        return redirect('contracts:view_acknowledgement_letter', clin_id=clin.id)
    
    context = {
        'clin': clin,
        'existing_letter': existing_letter,
    }
    return render(request, 'contracts/generate_acknowledgement_letter.html', context)

@conditional_login_required
def view_acknowledgement_letter(request, clin_id):
    clin = get_object_or_404(Clin, id=clin_id)
    letter = get_object_or_404(AcknowledgementLetter, clin=clin)
    
    context = {
        'clin': clin,
        'letter': letter,
    }
    return render(request, 'contracts/view_acknowledgement_letter.html', context)

@method_decorator(conditional_login_required, name='dispatch')
class AcknowledgementLetterUpdateView(UpdateView):
    model = AcknowledgementLetter
    template_name = 'contracts/acknowledgement_letter_form.html'
    fields = [
        'salutation', 'addr_fname', 'addr_lname', 'supplier',
        'st_address', 'city', 'state', 'zip', 'po', 'po_ext',
        'contract_num', 'fat_plt_due_date', 'supplier_due_date',
        'dpas_priority', 'statz_contact', 'statz_contact_title',
        'statz_contact_phone', 'statz_contact_email'
    ]
    
    def form_valid(self, form):
        form.instance.modified_by = self.request.user
        form.instance.letter_date = timezone.now()
        response = super().form_valid(form)
        messages.success(self.request, "Acknowledgement letter updated successfully.")
        return response
    
    def get_success_url(self):
        return reverse('contracts:view_acknowledgement_letter', kwargs={'clin_id': self.object.clin.id})

@conditional_login_required
def extract_dd1155_data(request):
    """
    View to handle DD Form 1155 file upload and data extraction using coordinate-based approach
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Only POST requests are allowed'})
    
    if 'dd1155_file' not in request.FILES:
        return JsonResponse({'success': False, 'error': 'No file was uploaded'})
    
    uploaded_file = request.FILES['dd1155_file']
    
    # Check if the file is a PDF
    if not uploaded_file.name.lower().endswith('.pdf'):
        return JsonResponse({'success': False, 'error': 'Uploaded file must be a PDF'})
    
    # Create a temporary file to store the uploaded PDF
    temp_file = None
    try:
        # Create a temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        temp_file.write(uploaded_file.read())
        temp_file.close()
        
        # Extract text from the PDF using coordinate-based approach
        extraction_results = extract_text_from_pdf(temp_file.name)
        
        # Parse the extracted text to get contract data
        contract_data = parse_dd1155_text(extraction_results)
        
        # Include the raw extraction results for debugging
        if isinstance(extraction_results, dict):
            # Include coordinate results for debugging
            contract_data['coordinate_results'] = extraction_results.get('coordinate_results', {})
            # Include the raw text in the response
            contract_data['raw_text'] = extraction_results.get('full_text', '')
        else:
            # Fallback if extraction_results is not a dict
            contract_data['raw_text'] = str(extraction_results)
        
        return JsonResponse({
            'success': True,
            **contract_data
        })
    
    except Exception as e:
        logger.error(f"Error processing DD Form 1155: {str(e)}")
        return JsonResponse({'success': False, 'error': f'Error processing file: {str(e)}'})
    
    finally:
        # Clean up the temporary file
        if temp_file and os.path.exists(temp_file.name):
            os.unlink(temp_file.name)

def extract_text_from_pdf(pdf_path):
    """
    Extract text from a PDF file using PyMuPDF's coordinate-based approach.
    Falls back to PyPDF2 and OCR if needed.
    """
    # Define the box areas for each field as percentages (left, upper, right, lower)
    coordinates = {
        'contract_number': (73 / 1224, 141 / 1584, 359 / 1224, 174 / 1584),
        'award_date': (592 / 1224, 157 / 1584, 760 / 1224, 174 / 1584),
        'buyer': (73 / 1224, 201 / 1584, 428 / 1224, 283 / 1584),
        'po_number': (787 / 1224, 141 / 1584, 1002 / 1224, 174 / 1584),
        'contract_type_purchase': (201 / 1224, 576 / 1584, 228 / 1224, 621 / 1584),
        'contract_type_delivery': (201 / 1224, 541 / 1584, 228 / 1224, 571 / 1584),
        'due_date_days': (787 / 1224, 315 / 1584, 1002 / 1224, 334 / 1584),
        'contract_amount': (1002 / 1224, 1054 / 1584, 1150 / 1224, 1075 / 1584)
    }
    
    # Initialize results dictionary
    results = {key: 'Not found' for key in coordinates.keys()}
    
    # Try to extract text using PyMuPDF's coordinate-based approach
    try:
        # Open the PDF document
        pdf_document = fitz.open(pdf_path)
        
        # Extract text from the specific box areas of the first page
        if pdf_document.page_count > 0:
            first_page = pdf_document.load_page(0)
            width, height = first_page.rect.width, first_page.rect.height
            
            # Convert percentages to coordinates based on the page size
            new_coordinates = {key: (
                left_pct * width,
                upper_pct * height,
                right_pct * width,
                lower_pct * height
            ) for key, (left_pct, upper_pct, right_pct, lower_pct) in coordinates.items()}
            
            # Extract text from each box
            for key, box in new_coordinates.items():
                text = first_page.get_textbox(box).strip()
                results[key] = text if text else 'Not found'
            
            # Get full text for fallback
            full_text = ""
            for page_num in range(pdf_document.page_count):
                page = pdf_document.load_page(page_num)
                full_text += page.get_text()
            
            # Close the document
            pdf_document.close()
            
            # Return both the coordinate-based results and the full text
            return {
                'coordinate_results': results,
                'full_text': full_text
            }
            
    except Exception as e:
        logger.error(f"Error extracting text with PyMuPDF: {str(e)}")
    
    # Fallback to PyPDF2 if PyMuPDF fails
    full_text = ""
    try:
        with open(pdf_path, 'rb') as file:
            reader = PyPDF2.PdfReader(file, strict=False)
            
            for page in range(len(reader.pages)):
                full_text += reader.pages[page].extract_text() + "\n"
    except Exception as e:
        logger.error(f"Error extracting text with PyPDF2: {str(e)}")
    
    # If we got meaningful text, return it
    if len(full_text.strip()) > 100:  # Assuming a form should have more than 100 chars of text
        return {
            'coordinate_results': {'error': 'PyMuPDF extraction failed'},
            'full_text': full_text
        }
    
    # If PyPDF2 didn't extract enough text, try OCR
    try:
        # Check if pytesseract is properly configured
        try:
            pytesseract.get_tesseract_version()
        except Exception as e:
            logger.error(f"Tesseract OCR not properly configured: {str(e)}")
            return {
                'coordinate_results': {'error': 'PyMuPDF extraction failed'},
                'full_text': full_text
            }
        
        # Convert PDF to images
        images = pdf2image.convert_from_path(pdf_path)
        
        # Perform OCR on each image
        for img in images:
            full_text += pytesseract.image_to_string(img) + "\n"
            
    except Exception as e:
        logger.error(f"Error performing OCR: {str(e)}")
    
    return {
        'coordinate_results': {'error': 'PyMuPDF extraction failed'},
        'full_text': full_text
    }

def parse_dd1155_text(extraction_results):
    """
    Parse text extracted from DD Form 1155 to get contract information
    using both coordinate-based results and full text extraction
    """
    # Initialize data dictionary
    data = {
        'contract_number': None,
        'po_number': None,
        'award_date': None,
        'due_date': None,
        'contract_type': None,
        'buyer': None,
        'contract_amount': None
    }
    
    # If we have coordinate results, use them as primary source
    if isinstance(extraction_results, dict) and 'coordinate_results' in extraction_results:
        coord_results = extraction_results['coordinate_results']
        full_text = extraction_results.get('full_text', '')
        
        # Process contract number
        if 'contract_number' in coord_results and coord_results['contract_number'] != 'Not found':
            data['contract_number'] = coord_results['contract_number'].strip()
        
        # Process PO number
        if 'po_number' in coord_results and coord_results['po_number'] != 'Not found':
            data['po_number'] = coord_results['po_number'].strip()
        
        # Process award date
        if 'award_date' in coord_results and coord_results['award_date'] != 'Not found':
            date_str = coord_results['award_date'].strip()
            try:
                # Try to parse the date
                date_formats = ['%m/%d/%Y', '%m-%d-%Y', '%m/%d/%y', '%m-%d-%y']
                for fmt in date_formats:
                    try:
                        parsed_date = datetime.strptime(date_str, fmt)
                        data['award_date'] = parsed_date.strftime('%Y-%m-%d')
                        break
                    except ValueError:
                        continue
            except Exception as e:
                logger.warning(f"Error parsing award date: {e}")
        
        # Process due date
        if 'due_date_days' in coord_results and coord_results['due_date_days'] != 'Not found':
            due_date_text = coord_results['due_date_days'].strip()
            try:
                # Check if it's a number of days
                if re.search(r'\d+\s*DAYS', due_date_text, re.IGNORECASE):
                    days_match = re.search(r'(\d+)\s*DAYS', due_date_text, re.IGNORECASE)
                    if days_match and data['award_date']:
                        days = int(days_match.group(1))
                        award_date = datetime.strptime(data['award_date'], '%Y-%m-%d')
                        due_date = award_date + timedelta(days=days)
                        data['due_date'] = due_date.strftime('%Y-%m-%d')
                else:
                    # Try to parse as a direct date
                    date_formats = ['%m/%d/%Y', '%m-%d-%Y', '%m/%d/%y', '%m-%d-%y']
                    for fmt in date_formats:
                        try:
                            parsed_date = datetime.strptime(due_date_text, fmt)
                            data['due_date'] = parsed_date.strftime('%Y-%m-%d')
                            break
                        except ValueError:
                            continue
            except Exception as e:
                logger.warning(f"Error parsing due date: {e}")
        
        # Process contract type
        contract_type = None
        if 'contract_type_purchase' in coord_results and coord_results['contract_type_purchase'] != 'Not found':
            if 'X' in coord_results['contract_type_purchase']:
                contract_type = 'Purchase Order'
        
        if not contract_type and 'contract_type_delivery' in coord_results and coord_results['contract_type_delivery'] != 'Not found':
            if 'X' in coord_results['contract_type_delivery']:
                contract_type = 'Delivery Order'
        
        if contract_type:
            data['contract_type'] = contract_type
        
        # Process buyer information
        if 'buyer' in coord_results and coord_results['buyer'] != 'Not found':
            data['buyer'] = coord_results['buyer'].strip()
            # Limit buyer name to reasonable length
            if len(data['buyer']) > 50:
                data['buyer'] = data['buyer'][:50]
        
        # Process contract amount
        if 'contract_amount' in coord_results and coord_results['contract_amount'] != 'Not found':
            amount_str = coord_results['contract_amount'].strip()
            # Remove any non-numeric characters except decimal point
            amount_str = re.sub(r'[^\d.]', '', amount_str)
            try:
                if amount_str:
                    data['contract_amount'] = amount_str
            except Exception as e:
                logger.warning(f"Error parsing contract amount: {e}")
        
        # If we have full text, use it as a fallback for missing data
        if full_text and (not data['contract_number'] or not data['po_number'] or 
                          not data['award_date'] or not data['due_date'] or 
                          not data['contract_type'] or not data['buyer']):
            fallback_data = parse_full_text(full_text)
            
            # Fill in missing data from fallback
            for key, value in fallback_data.items():
                if not data[key] and value:
                    data[key] = value
    else:
        # If we don't have coordinate results, treat the input as full text
        data = parse_full_text(extraction_results)
    
    return data

def parse_full_text(text):
    """
    Parse full text extracted from DD Form 1155 using regex patterns
    """
    data = {
        'contract_number': None,
        'po_number': None,
        'award_date': None,
        'due_date': None,
        'contract_type': None,
        'buyer': None,
        'contract_amount': None
    }
    
    # Contract Number (usually in format like "N00039-19-F-0001")
    contract_number_patterns = [
        r'[A-Z0-9]{6}-[0-9]{2}-[A-Z0-9]{1}-[0-9]{4}',
        r'CONTRACT\s*NO\.\s*([\w\-]+)',
        r'CONTRACT\s*NUMBER\s*:\s*([\w\-]+)',
        r'CONTRACT\s*#\s*:\s*([\w\-]+)',
        r'ORDER\s*NUMBER\s*:\s*([\w\-]+)',
        r'ORDER\s*NO\.\s*([\w\-]+)'
    ]
    
    for pattern in contract_number_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            if match.groups():
                data['contract_number'] = match.group(1).strip()
            else:
                data['contract_number'] = match.group(0).strip()
            break
    
    # PO Number
    po_patterns = [
        r'PURCHASE\s*ORDER\s*NO\.\s*([\w\-]+)',
        r'PURCHASE\s*ORDER\s*NUMBER\s*:\s*([\w\-]+)',
        r'P\.?O\.?\s*NUMBER\s*:\s*([\w\-]+)',
        r'P\.?O\.?\s*NO\.\s*([\w\-]+)',
        r'REQUISITION\s*NUMBER\s*:\s*([\w\-]+)',
        r'REQUISITION\s*NO\.\s*([\w\-]+)'
    ]
    
    for pattern in po_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            data['po_number'] = match.group(1).strip()
            break
    
    # Award Date
    award_date_patterns = [
        r'DATE\s*OF\s*ORDER\s*:\s*(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})',
        r'ORDER\s*DATE\s*:\s*(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})',
        r'AWARD\s*DATE\s*:\s*(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})',
        r'DATE\s*ISSUED\s*:\s*(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})'
    ]
    
    for pattern in award_date_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            date_str = match.group(1).strip()
            # Try to parse and standardize the date format
            try:
                # Handle different date formats
                if '/' in date_str:
                    parts = date_str.split('/')
                elif '-' in date_str:
                    parts = date_str.split('-')
                else:
                    continue
                
                # Assuming MM/DD/YYYY or MM-DD-YYYY format
                if len(parts) == 3:
                    month, day, year = parts
                    # Handle 2-digit years
                    if len(year) == 2:
                        year = '20' + year if int(year) < 50 else '19' + year
                    data['award_date'] = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
            except Exception as e:
                logger.warning(f"Error parsing award date: {e}")
            break
    
    # Due Date / Delivery Date
    due_date_patterns = [
        r'DELIVERY\s*DATE\s*:\s*(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})',
        r'REQUIRED\s*DELIVERY\s*DATE\s*:\s*(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})',
        r'DELIVERY\s*BY\s*:\s*(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})',
        r'DUE\s*DATE\s*:\s*(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})'
    ]
    
    for pattern in due_date_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            date_str = match.group(1).strip()
            # Try to parse and standardize the date format
            try:
                # Handle different date formats
                if '/' in date_str:
                    parts = date_str.split('/')
                elif '-' in date_str:
                    parts = date_str.split('-')
                else:
                    continue
                
                # Assuming MM/DD/YYYY or MM-DD-YYYY format
                if len(parts) == 3:
                    month, day, year = parts
                    # Handle 2-digit years
                    if len(year) == 2:
                        year = '20' + year if int(year) < 50 else '19' + year
                    data['due_date'] = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
            except Exception as e:
                logger.warning(f"Error parsing due date: {e}")
            break
    
    # Contract Type
    contract_type_patterns = [
        r'CONTRACT\s*TYPE\s*:\s*(\w+)',
        r'TYPE\s*OF\s*ORDER\s*:\s*(\w+)',
        r'ORDER\s*TYPE\s*:\s*(\w+)'
    ]
    
    contract_type_keywords = {
        'FFP': 'Firm Fixed Price',
        'FIRM FIXED PRICE': 'Firm Fixed Price',
        'FIXED PRICE': 'Firm Fixed Price',
        'CPFF': 'Cost Plus Fixed Fee',
        'COST PLUS FIXED FEE': 'Cost Plus Fixed Fee',
        'T&M': 'Time and Materials',
        'TIME AND MATERIALS': 'Time and Materials',
        'IDIQ': 'Indefinite Delivery Indefinite Quantity',
        'INDEFINITE DELIVERY': 'Indefinite Delivery Indefinite Quantity',
        'PURCHASE ORDER': 'Purchase Order',
        'DELIVERY ORDER': 'Delivery Order'
    }
    
    for pattern in contract_type_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            contract_type = match.group(1).strip().upper()
            # Map abbreviated contract types to full names
            if contract_type in contract_type_keywords:
                data['contract_type'] = contract_type_keywords[contract_type]
            else:
                data['contract_type'] = contract_type
            break
    
    # If contract type not found by pattern, search for keywords
    if not data['contract_type']:
        for keyword, full_name in contract_type_keywords.items():
            if re.search(r'\b' + re.escape(keyword) + r'\b', text, re.IGNORECASE):
                data['contract_type'] = full_name
                break
    
    # Buyer Information
    buyer_patterns = [
        r'BUYER\s*:\s*([A-Za-z\s\.]+)',
        r'CONTRACTING\s*OFFICER\s*:\s*([A-Za-z\s\.]+)',
        r'ISSUED\s*BY\s*:\s*([A-Za-z\s\.]+)'
    ]
    
    for pattern in buyer_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            data['buyer'] = match.group(1).strip()
            # Limit buyer name to reasonable length
            if len(data['buyer']) > 50:
                data['buyer'] = data['buyer'][:50]
            break
    
    # Contract Amount
    amount_patterns = [
        r'TOTAL\s*AMOUNT\s*:\s*\$?(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)',
        r'TOTAL\s*:\s*\$?(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)',
        r'AMOUNT\s*:\s*\$?(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)'
    ]
    
    for pattern in amount_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            amount_str = match.group(1).strip()
            # Remove commas
            amount_str = amount_str.replace(',', '')
            data['contract_amount'] = amount_str
            break
    
    return data

@method_decorator(conditional_login_required, name='dispatch')
class ContractLogView(ListView):
    model = Contract
    template_name = 'contracts/contract_log_view.html'
    context_object_name = 'contracts'
    paginate_by = 50
    
    def get_queryset(self):
        queryset = Contract.objects.all().order_by('-award_date')
        
        # Filter by sales class if provided
        sales_class = self.request.GET.get('sales_class')
        if sales_class and sales_class != 'all':
            queryset = queryset.filter(sales_class=sales_class)
            
        # Filter by search term if provided
        search_term = self.request.GET.get('search')
        if search_term:
            queryset = queryset.filter(
                Q(contract_number__icontains=search_term) |
                Q(title__icontains=search_term) |
                Q(customer__icontains=search_term) |
                Q(sales_class__icontains=search_term)
            )
            
        return queryset
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get unique sales classes for filter dropdown
        context['sales_classes'] = Contract.objects.values_list(
            'sales_class', flat=True).distinct().order_by('sales_class')
        
        # Add today's date for reference
        context['today'] = timezone.now().date()
        
        # Add export functionality status
        context['export_available'] = OPENPYXL_AVAILABLE
        
        # Get export folder size if available
        try:
            export_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
            if not os.path.exists(export_dir):
                os.makedirs(export_dir)
            context['export_folder_size'] = len(os.listdir(export_dir))
        except Exception:
            context['export_folder_size'] = 0
            
        return context

@conditional_login_required
def export_contract_log(request):
    if not OPENPYXL_AVAILABLE:
        messages.error(request, "Export functionality is not available. Please install openpyxl package.")
        return redirect('contracts:contract_log_view')
        
    if request.method != 'POST':
        messages.error(request, "Invalid request method.")
        return redirect('contracts:contract_log_view')
    
    # Get export format and filename
    export_format = request.POST.get('export_format', 'excel')
    filename = request.POST.get('filename', f'contract_log_{timezone.now().strftime("%Y%m%d")}')
    
    # Filter contracts based on criteria
    queryset = Contract.objects.all().order_by('-award_date')
    
    sales_class = request.POST.get('sales_class')
    if sales_class and sales_class != 'all':
        queryset = queryset.filter(sales_class=sales_class)
        
    search_term = request.POST.get('search')
    if search_term:
        queryset = queryset.filter(
            Q(contract_number__icontains=search_term) |
            Q(title__icontains=search_term) |
            Q(customer__icontains=search_term) |
            Q(sales_class__icontains=search_term)
        )
    
    # Create exports directory if it doesn't exist
    export_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
    if not os.path.exists(export_dir):
        os.makedirs(export_dir)
    
    # Export based on format
    if export_format == 'excel' and OPENPYXL_AVAILABLE:
        # Export to Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contract Log"
        
        # Add headers
        headers = [
            'Contract Number', 'Title', 'Customer', 'Sales Class', 
            'Award Date', 'Period of Performance', 'Value', 'Status'
        ]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        # Add data
        for row_num, contract in enumerate(queryset, 2):
            # Determine status
            if contract.cancelled:
                status = "Cancelled"
            elif not contract.open:
                status = "Closed"
            else:
                status = "Open"
                
            # Format period of performance
            pop = f"{contract.pop_start.strftime('%m/%d/%Y')} - {contract.pop_end.strftime('%m/%d/%Y')}" if contract.pop_start and contract.pop_end else "N/A"
            
            # Add row data
            row_data = [
                contract.contract_number,
                contract.title,
                contract.customer,
                contract.sales_class,
                contract.award_date.strftime('%m/%d/%Y') if contract.award_date else "N/A",
                pop,
                f"${contract.value:,.2f}" if contract.value else "$0.00",
                status
            ]
            
            for col_num, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=cell_value)
        
        # Save the workbook
        file_path = os.path.join(export_dir, f"{filename}.xlsx")
        wb.save(file_path)
        
        messages.success(request, f"Excel export created successfully: {filename}.xlsx")
        
    elif export_format == 'csv':
        # Export to CSV
        file_path = os.path.join(export_dir, f"{filename}.csv")
        
        with open(file_path, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            
            # Write headers
            writer.writerow([
                'Contract Number', 'Title', 'Customer', 'Sales Class', 
                'Award Date', 'Period of Performance', 'Value', 'Status'
            ])
            
            # Write data
            for contract in queryset:
                # Determine status
                if contract.cancelled:
                    status = "Cancelled"
                elif not contract.open:
                    status = "Closed"
                else:
                    status = "Open"
                    
                # Format period of performance
                pop = f"{contract.pop_start.strftime('%m/%d/%Y')} - {contract.pop_end.strftime('%m/%d/%Y')}" if contract.pop_start and contract.pop_end else "N/A"
                
                # Write row
                writer.writerow([
                    contract.contract_number,
                    contract.title,
                    contract.customer,
                    contract.sales_class,
                    contract.award_date.strftime('%m/%d/%Y') if contract.award_date else "N/A",
                    pop,
                    f"${contract.value:,.2f}" if contract.value else "$0.00",
                    status
                ])
        
        messages.success(request, f"CSV export created successfully: {filename}.csv")
        
    elif export_format == 'pdf':
        # Simple PDF export (as text file with .pdf extension)
        file_path = os.path.join(export_dir, f"{filename}.pdf")
        
        with open(file_path, 'w') as f:
            f.write("Contract Log\n\n")
            
            # Write headers
            f.write("Contract Number\tTitle\tCustomer\tSales Class\tAward Date\tPeriod of Performance\tValue\tStatus\n")
            
            # Write data
            for contract in queryset:
                # Determine status
                if contract.cancelled:
                    status = "Cancelled"
                elif not contract.open:
                    status = "Closed"
                else:
                    status = "Open"
                    
                # Format period of performance
                pop = f"{contract.pop_start.strftime('%m/%d/%Y')} - {contract.pop_end.strftime('%m/%d/%Y')}" if contract.pop_start and contract.pop_end else "N/A"
                
                # Write row
                f.write(f"{contract.contract_number}\t{contract.title}\t{contract.customer}\t{contract.sales_class}\t")
                f.write(f"{contract.award_date.strftime('%m/%d/%Y') if contract.award_date else 'N/A'}\t{pop}\t")
                f.write(f"${contract.value:,.2f}" if contract.value else "$0.00")
                f.write(f"\t{status}\n")
        
        messages.success(request, f"PDF export created successfully: {filename}.pdf")
    
    else:
        messages.error(request, f"Unsupported export format: {export_format}")
    
    return redirect('contracts:contract_log_view')

@conditional_login_required
def open_export_folder(request):
    export_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
    
    # Create directory if it doesn't exist
    if not os.path.exists(export_dir):
        os.makedirs(export_dir)
    
    # Open folder based on OS
    try:
        if sys.platform == 'win32':
            os.startfile(export_dir)
        elif sys.platform == 'darwin':  # macOS
            subprocess.run(['open', export_dir])
        else:  # Linux
            subprocess.run(['xdg-open', export_dir])
        messages.success(request, "Export folder opened successfully.")
    except Exception as e:
        messages.error(request, f"Failed to open export folder: {str(e)}")
    
    return redirect('contracts:contract_log_view')
