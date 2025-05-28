"""
Views for the reports app.

This module provides views for handling natural language queries and displaying results.
"""
import logging
import time
import json
from typing import Dict, Any
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.generic import TemplateView, ListView, DetailView, CreateView, UpdateView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.core.paginator import Paginator
from django.core.exceptions import FieldDoesNotExist
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from django.views.generic import View
from django.db.models import Count
from django.utils import timezone
import re
from django.urls import reverse, reverse_lazy
from django.contrib import messages
from django.db import connection
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.conf import settings
import requests
from .models import ReportRequest, Report, ReportChange
from django.apps import apps
from contracts.utils.contracts_schema import generate_contracts_schema_description, generate_condensed_contracts_schema
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator


class UserReportListView(LoginRequiredMixin, ListView):
    """View for users to see their report requests and completed reports."""
    
    model = ReportRequest
    template_name = 'reports/user_reports.html'
    context_object_name = 'reports'
    
    def get_queryset(self):
        return ReportRequest.objects.filter(user=self.request.user, status__in=['in_progress', 'pending']).order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['completed_reports'] = Report.objects.filter(
            report_request__user=self.request.user
        ).select_related('report_request')
        return context

class ReportRequestCreateView(LoginRequiredMixin, CreateView):
    """View for users to create new report requests."""
    
    model = ReportRequest
    fields = ['request_text']
    template_name = 'reports/report_request_form.html'
    success_url = reverse_lazy('reports:user-reports')
    
    def form_valid(self, form):
        form.instance.user = self.request.user
        response = super().form_valid(form)
        
        from users.models import SystemMessage
        
        # Get the user from the email address in settings
        try:
            # Create a message for the report creator
            SystemMessage.create_message(
                user=User.objects.get(email=settings.REPORT_CREATOR_EMAIL),
                title="Report Request Submitted",
                message=f"A new report request '{self.object.generated_name}' has been submitted.",
                priority="medium",
                source_app="reports",
                source_model="ReportRequest",
                source_id=str(self.object.id),
                action_url=reverse('reports:creator-reports')
            )
            messages.success(self.request, 'Your report request has been submitted successfully.')
            return response
        except User.DoesNotExist:
            messages.error(self.request, f'Report creator with email {settings.REPORT_CREATOR_EMAIL} not found. Please contact your administrator.')
            return redirect('reports:user-reports')

class ReportCreatorListView(UserPassesTestMixin, ListView):
    """View for report creators to see all report requests."""
    
    model = ReportRequest
    template_name = 'reports/creator_reports.html'
    context_object_name = 'reports'
    
    def test_func(self):
        return self.request.user.is_staff
    
    def get_queryset(self):
        status = self.request.GET.get('status', '')
        queryset = ReportRequest.objects.all().select_related('user', 'assigned_to')
        
        if status:
            queryset = queryset.filter(status=status)
            
        return queryset.order_by('-created_at')

class ReportCreatorDetailView(UserPassesTestMixin, UpdateView):
    """View for report creators to update report requests and add SQL."""
    
    model = ReportRequest
    template_name = 'reports/creator_detail.html'
    fields = ['status', 'assigned_to']
    
    def test_func(self):
        return self.request.user.is_staff
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['report_form'] = self.get_report_form()
        return context
    
    def get_report_form(self):
        report = getattr(self.object, 'completed_report', None)
        if report:
            return {'sql_query': report.sql_query, 'description': report.description}
        return {'sql_query': '', 'description': ''}
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        
        # Handle report creation/update
        if 'sql_query' in request.POST:
            report, created = Report.objects.update_or_create(
                report_request=self.object,
                defaults={
                    'sql_query': request.POST['sql_query'],
                    'description': request.POST.get('description', '')
                }
            )
            
            if created:
                messages.success(request, 'Report has been created successfully.')
            else:
                messages.success(request, 'Report has been updated successfully.')
                
            return redirect('reports:creator-detail', pk=self.object.pk)
        
        return super().post(request, *args, **kwargs)

class ReportViewView(LoginRequiredMixin, DetailView):
    """View for users to view their completed reports."""
    
    model = Report
    template_name = 'reports/report_view.html'
    
    def get_queryset(self):
        return Report.objects.filter(report_request__user=self.request.user)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['change_requests'] = self.object.change_requests.all().order_by('-created_at')
        
        # Get page size from request or user settings
        page_size = self.request.GET.get('page_size', '25')
        page_number = self.request.GET.get('page', 1)
        
        # Execute the SQL query and get results
        try:
            with connection.cursor() as cursor:
                cursor.execute(self.object.sql_query)
                context['columns'] = [col[0] for col in cursor.description]
                all_results = cursor.fetchall()
                
                # Handle pagination
                if page_size.upper() != 'ALL':
                    paginator = Paginator(list(all_results), int(page_size))
                    page_obj = paginator.get_page(page_number)
                    context['results'] = page_obj.object_list
                    context['page_obj'] = page_obj
                else:
                    context['results'] = all_results
                
                context['total_records'] = len(all_results)
                context['page_size'] = page_size
                context['available_page_sizes'] = ['25', '50', '100', 'ALL']
                
        except Exception as e:
            context['error'] = str(e)
            context['columns'] = []
            context['results'] = []
            logger = logging.getLogger(__name__)
            logger.error(f"Error executing report query for report {self.object.pk}: {str(e)}")
            
        return context

class ReportChangeCreateView(LoginRequiredMixin, CreateView):
    """View for users to request changes to a report."""
    
    model = ReportChange
    fields = ['change_text']
    template_name = 'reports/report_change_form.html'
    
    def get_report(self):
        return get_object_or_404(Report, pk=self.kwargs['report_pk'], report_request__user=self.request.user)
    
    def form_valid(self, form):
        report = self.get_report()
        form.instance.report = report
        form.instance.user = self.request.user
        response = super().form_valid(form)
        messages.success(self.request, 'Your change request has been submitted.')
        return response
    
    def get_success_url(self):
        return reverse_lazy('reports:report-view', kwargs={'pk': self.kwargs['report_pk']})

@login_required
def export_report(request, pk):
    """Export a report to Excel."""
    report = get_object_or_404(Report, pk=pk, report_request__user=request.user)
    
    # Create a new workbook and select the active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report Data"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add report metadata
    ws.merge_cells('A1:D1')
    ws['A1'] = report.report_request.generated_name
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:D2')
    ws['A2'] = report.description or "No description provided"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A3'].font = Font(italic=True)
    
    # Add a blank row
    current_row = 5
    
    try:
        # Execute the SQL query and get results
        with connection.cursor() as cursor:
            cursor.execute(report.sql_query)
            columns = [col[0] for col in cursor.description]
            results = cursor.fetchall()
        
        # Write headers
        for col_num, column in enumerate(columns, 1):
            cell = ws.cell(row=current_row, column=col_num, value=column)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            
            # Write data
        for row in results:
            current_row += 1
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=current_row, column=col_num, value=value)
                cell.border = border
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
    except Exception as e:
        # If there's an error executing the query, create an error sheet
        ws['A5'] = "Error executing report"
        ws['A6'] = str(e)
        logger = logging.getLogger(__name__)
        logger.error(f"Error exporting report {pk}: {str(e)}")
    
    # Create the response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    response['Content-Disposition'] = f'attachment; filename="{report.report_request.generated_name}.xlsx"'
    
    # Save the workbook to the response
    wb.save(response)
    
    return response

def ai_easy_call(request):
    if request.method == 'POST':
        user_query = request.POST.get('query') # This will be something like "what day is today?"

        # SLIMMED DOWN PROMPT FOR TESTING CONNECTION
        prompt = f"""
        Answer the following question directly and concisely.

        User's question: "{user_query}"
        """

        api_key = "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"  # Replace with your actual OpenRouter API key
        ai_url = "https://openrouter.ai/api/v1/chat/completions"
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }
        data = {
            "model": "mistralai/devstral-small:free",  # Or another free model from OpenRouter
            "messages": [
                {"role": "system", "content": "You are a helpful assistant."},
                {"role": "user", "content": prompt}
            ]
        }

        try:
            response = requests.post(ai_url, headers=headers, data=json.dumps(data))
            response.raise_for_status()

            # OpenRouter returns the response in choices[0]['message']['content']
            llm_response_text = response.json()['choices'][0]['message']['content']

            context = {
                'query': user_query,
                'generated_report': llm_response_text, # Just display the LLM's raw response
                'generated_sql': "N/A for simple query", # Indicate SQL is not expected
                'error': None
            }
        except requests.exceptions.ConnectionError:
            context = {'error': "Could not connect to OpenRouter.ai. Is your internet connection working?"}
        except requests.exceptions.RequestException as e:
            # Print error details for debugging
            error_text = getattr(e.response, 'text', str(e))
            context = {'error': f"Error communicating with OpenRouter.ai: {e}\n{error_text}"}
        except Exception as e:
            context = {'error': f"An unexpected error occurred: {e}"}

        return render(request, 'reports/report_results.html', context)
    
    return render(request, 'reports/generate_report_form.html')

def ai_generate_report_view(request):
    if request.method == 'POST':
        user_query = request.POST.get('query')
        is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'
        database_schema = generate_contracts_schema_description()
        prompt = f"""
Given the following database schema:
{database_schema}

User's request: \"{user_query}\"

Generate:
1. A concise, human-friendly report name (max 8 words, title case, no punctuation except dashes or spaces).
2. A valid SQL Server SQL query to answer the user's request. Select queries only. Anyone asking for an Update, Delete, or Insert query should be rejected.

Your output should strictly follow this format:
REPORT_NAME:
<your generated report name here>
SQL_QUERY:
```sql
-- Your generated SQL query here
```
        """
        api_key = "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"
        try:
            ai_url = "https://openrouter.ai/api/v1/chat/completions"
            headers = {
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json"
            }
            data = {
                "model": "mistralai/devstral-small:free",
                "messages": [
                    {"role": "system", "content": "You are an AI assistant that can generate SQL queries and concise report names from the user's input."},
                    {"role": "user", "content": prompt}
                ]
            }
            response = requests.post(ai_url, headers=headers, data=json.dumps(data))
            response.raise_for_status()
            llm_response_text = response.json()['choices'][0]['message']['content']

            # Parse report name
            report_name = None
            sql_start = llm_response_text.find('SQL_QUERY:')
            if llm_response_text.startswith('REPORT_NAME:'):
                # AI followed the format
                name_section = llm_response_text[len('REPORT_NAME:'):sql_start].strip() if sql_start != -1 else llm_response_text[len('REPORT_NAME:'):].strip()
                report_name = name_section.split('\n')[0].strip()
            else:
                # Fallback: try to find the first line as the name
                report_name = llm_response_text.split('\n')[0].strip()

            # Parse SQL
            sql_code_start = llm_response_text.find('```sql')
            sql_code_end = llm_response_text.find('```', sql_code_start + 1)
            generated_sql = ""
            if sql_code_start != -1 and sql_code_end != -1:
                generated_sql = llm_response_text[sql_code_start + len('```sql'):sql_code_end].strip()

            from .models import ReportRequest, Report
            user = request.user
            report_request = ReportRequest.objects.create(
                user=user,
                request_text=user_query,
                status='completed',
            )
            # Override generated_name if AI provided one
            if report_name:
                report_request.generated_name = report_name[:100]
                report_request.save(update_fields=['generated_name'])
            report = Report.objects.create(
                report_request=report_request,
                sql_query=generated_sql,
                description=f"AI-generated report for: {user_query}"
            )
            if is_ajax:
                return JsonResponse({
                    'success': True,
                    'report_id': str(report.id),
                    'report_request_id': str(report_request.id),
                    'generated_name': report_request.generated_name,
                    'redirect_url': reverse('reports:report-view', kwargs={'pk': report.id})
                })
            else:
                return redirect('reports:user-reports')
        except Exception as e:
            if is_ajax:
                return JsonResponse({'success': False, 'error': str(e)}, status=500)
            else:
                return render(request, 'reports/report_results.html', {'error': str(e)})
    return render(request, 'reports/generate_report_form.html')
